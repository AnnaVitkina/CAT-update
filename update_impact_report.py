"""
Build an Excel **impact preview** for an update CSV against a **prior** rate card.

**Default:** run with no arguments — you pick ``input/update/*.csv`` and ``processing/rate/*.json``
from the same numbered lists as ``update.py``; the CSV is combined in memory (no full merge).

Output workbook has **two sheets** (same columns):
  - **By update window** — one row per fee window; cost / classification cells are **multi-line** (previous style).
  - **By cost** — one row per projected cost (split rows).
  Columns: KEY, Rate Card, Carrier, SERVICE, Service, Valid from, Valid to,
  Origin/Destination, matching lanes, Cost to add, Validity of cost to add, Costs to change.

Uses the same lane routing and ``build_lane_costs_from_update`` scaffolding as ``update.py``
so cost names match the real merge.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from transform_inputs import ROOT, safe_json_name

from update import (
    build_lane_costs_from_update,
    combine_ocean_rates_csv,
    find_donor_lane,
    flat_records_from_combined_merge_rec,
    lanes_matching_route,
    prompt_pick_csv,
    prompt_pick_rate_json,
    service_cy_cy,
    _fmt_dmy,
    _norm,
    _parse_us_date,
    _price_equal,
    _tier_validity_bounds,
)


OUTPUT_DIR_DEFAULT = ROOT / "processing" / "output"

HEADERS = [
    "KEY",
    "Rate Card",
    "Carrier",
    "SERVICE",
    "Service",
    "Valid from",
    "Valid to",
    "Origin Port",
    "Destination Port",
    "Lanes in previous rate card",
    "Validity of lanes",
    "Cost to add",
    "Validity of cost to add",
    "Costs to change",
]


def _load_json(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _fmt_lane_validity(ln: dict) -> str:
    vf = _norm(ln.get("Valid from"))
    vt = _norm(ln.get("Valid to"))
    if vf or vt:
        return f"{vf} – {vt}".strip()
    return ""


def _cost_line(c: dict) -> str:
    nm = _norm(c.get("Cost"))
    if not nm:
        return ""
    ccy = _norm(c.get("Currency"))
    pr = c.get("Price")
    if ccy:
        return f"{nm} | {ccy} {pr}"
    return f"{nm} | {pr}"


def _fmt_cost_validity_for_row(
    c: dict,
    csv_eff,
    csv_exp,
) -> str:
    """Prefer parsed ``Validity period`` on the projected cost row; else CSV window dates."""
    d1, d2 = _tier_validity_bounds(c)
    if d1 is not None and d2 is not None:
        return f"{_fmt_dmy(d1)} – {_fmt_dmy(d2)}"
    if d1 is not None and d2 is None:
        return f"from {_fmt_dmy(d1)} (open-ended)"
    if csv_eff and csv_exp:
        return f"{_fmt_dmy(csv_eff)} – {_fmt_dmy(csv_exp)}"
    return ""


def _classify_one_cost(nc: dict, route_lanes: list[dict]) -> str:
    """Single projected cost: new column vs update existing vs same value."""
    nm = _norm(nc.get("Cost"))
    if not nm:
        return ""

    if not route_lanes:
        return "No matching route on rate card — merge would append a new lane."

    want_ccy = _norm(nc.get("Currency"))
    want_pr = nc.get("Price")

    matches: list[tuple[str, bool]] = []
    for ln in route_lanes:
        lid = _norm(ln.get("Lane #"))
        for c in ln.get("Costs") or []:
            if _norm(c.get("Cost")) != nm:
                continue
            same = _norm(c.get("Currency")) == want_ccy and _price_equal(
                c.get("Price"), want_pr
            )
            matches.append((lid, same))

    if not matches:
        return f"New column: {nm}"
    if all(m[1] for m in matches):
        return f"Existing: {nm} (same currency/price on lane(s) — optional refresh)"
    return f"Update existing: {nm} (currency and/or price will change)"


def _build_common_base_row(
    flat: dict,
    rate_card_label: str,
    car: str,
    svc: str,
    route_lanes: list[dict],
    eff_d,
    exp_d,
) -> list:
    lane_nums = ", ".join(_norm(x.get("Lane #")) for x in route_lanes) or "(none)"
    lane_validity = (
        " | ".join(_fmt_lane_validity(x) for x in route_lanes) if route_lanes else ""
    )
    return [
        flat.get("KEY"),
        rate_card_label,
        car,
        flat.get("SERVICE__C"),
        svc,
        _fmt_dmy(eff_d) if eff_d else "",
        _fmt_dmy(exp_d) if exp_d else "",
        _norm(flat.get("ORIGIN_LOCATION_NAME__C")),
        _norm(flat.get("DESTINATION_LOCATION_NAME__C")),
        lane_nums,
        lane_validity,
    ]


def build_impact_rows_by_window(
    combined: dict,
    rate: dict,
    rate_card_label: str,
) -> list[list]:
    """One row per update fee window; costs and classifications grouped with newlines."""
    lanes = rate.get("sheets", {}).get("Rate card", {}).get("lanes") or []
    donor = find_donor_lane(lanes, None)

    rows_out: list[list] = []
    for rec in combined.get("records") or []:
        for flat in flat_records_from_combined_merge_rec(rec):
            car = _norm(flat.get("CARRIER"))
            svc = service_cy_cy(flat.get("SERVICE__C", ""))
            origin = _norm(flat.get("ORIGIN_LOCATION_NAME__C"))
            dest = _norm(flat.get("DESTINATION_LOCATION_NAME__C"))
            route_lanes = lanes_matching_route(lanes, car, svc, origin, dest)

            eff_d = _parse_us_date(_norm(flat.get("RATE_EFFECTIVE_DATE__C")))
            exp_d = _parse_us_date(_norm(flat.get("RATE_EXPIRATION_DATE__C")))

            issues: list[str] = []
            new_costs = build_lane_costs_from_update(flat, donor, lanes, issues)

            base_row = _build_common_base_row(
                flat, rate_card_label, car, svc, route_lanes, eff_d, exp_d
            )

            cost_to_add = "\n".join(
                x for x in (_cost_line(c) for c in new_costs) if x
            )

            if eff_d and exp_d:
                cost_validity = f"{_fmt_dmy(eff_d)} – {_fmt_dmy(exp_d)}"
            else:
                cost_validity = ""

            costs_change = "\n".join(
                _classify_one_cost(c, route_lanes)
                for c in new_costs
                if _norm(c.get("Cost"))
            )

            rows_out.append(
                base_row + [cost_to_add, cost_validity, costs_change]
            )
    return rows_out


def build_impact_rows_by_cost(
    combined: dict,
    rate: dict,
    rate_card_label: str,
) -> list[list]:
    """One row per projected cost row."""
    lanes = rate.get("sheets", {}).get("Rate card", {}).get("lanes") or []
    donor = find_donor_lane(lanes, None)

    rows_out: list[list] = []
    for rec in combined.get("records") or []:
        for flat in flat_records_from_combined_merge_rec(rec):
            car = _norm(flat.get("CARRIER"))
            svc = service_cy_cy(flat.get("SERVICE__C", ""))
            origin = _norm(flat.get("ORIGIN_LOCATION_NAME__C"))
            dest = _norm(flat.get("DESTINATION_LOCATION_NAME__C"))
            route_lanes = lanes_matching_route(lanes, car, svc, origin, dest)

            eff_d = _parse_us_date(_norm(flat.get("RATE_EFFECTIVE_DATE__C")))
            exp_d = _parse_us_date(_norm(flat.get("RATE_EXPIRATION_DATE__C")))

            issues: list[str] = []
            new_costs = build_lane_costs_from_update(flat, donor, lanes, issues)

            base_row = _build_common_base_row(
                flat, rate_card_label, car, svc, route_lanes, eff_d, exp_d
            )

            if not new_costs:
                continue

            for nc in new_costs:
                cline = _cost_line(nc)
                if not cline:
                    continue
                cost_validity = _fmt_cost_validity_for_row(nc, eff_d, exp_d)
                costs_change = _classify_one_cost(nc, route_lanes)
                rows_out.append(
                    base_row + [cline, cost_validity, costs_change]
                )
    return rows_out


def _fill_sheet(ws, rows: list[list]) -> None:
    header_font = Font(bold=True)
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [36, 28, 12, 14, 12, 12, 12, 14, 14, 22, 40, 50, 28, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_impact_workbook(
    path: Path,
    rows_by_window: list[list],
    rows_by_cost: list[list],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "By update window"
    _fill_sheet(ws1, rows_by_window)

    ws2 = wb.create_sheet("By cost")
    _fill_sheet(ws2, rows_by_cost)

    wb.save(path)


def main() -> None:
    p = argparse.ArgumentParser(
        description=(
            "Excel preview of update vs prior rate card (lanes + costs). "
            "Run with no args to choose CSV + rate JSON from menus (same as update.py). "
            "Does not run merge or write result JSON/Excel."
        )
    )
    p.add_argument(
        "-i",
        "--interactive",
        action="store_true",
        help=(
            "Force menu selection (same as default). Optional if you omit --update / --rate."
        ),
    )
    p.add_argument(
        "--update",
        type=Path,
        default=None,
        help=(
            "Optional: read combined JSON instead of choosing a CSV "
            "(e.g. processing/update_to_perform/…_combined.json). Requires --rate."
        ),
    )
    p.add_argument(
        "--rate",
        type=Path,
        default=None,
        help="Optional: prior rate JSON path. Required if --update is set.",
    )
    p.add_argument(
        "--out-dir",
        type=Path,
        default=OUTPUT_DIR_DEFAULT,
        help=f"Output folder (default: {OUTPUT_DIR_DEFAULT}).",
    )
    p.add_argument(
        "--rate-card-label",
        type=str,
        default="",
        help="Value for the Rate Card column (default: stem of --update file).",
    )
    p.add_argument(
        "--output-name",
        type=str,
        default="",
        help="Excel filename without path (default: update_impact_<safe_stem>.xlsx).",
    )
    args = p.parse_args()

    use_files = args.update is not None or args.rate is not None
    if use_files and (args.update is None or args.rate is None):
        p.error("With file mode, pass both --update and --rate, or pass neither to pick from menus.")

    # Default: pick from listed CSvs / rate JSONs. Use explicit paths only when both are given
    # and -i is not set.
    use_menus = bool(args.interactive) or not (use_files and not args.interactive)
    if use_menus:
        print("\nImpact report — choose files (same lists as update.py).\n")
        csv_path = prompt_pick_csv()
        if not csv_path or not csv_path.is_file():
            print("No CSV selected.")
            raise SystemExit(1)
        csv_path = csv_path.resolve()
        rate_path = prompt_pick_rate_json()
        if not rate_path or not rate_path.is_file():
            print("No rate JSON selected or file missing.")
            raise SystemExit(1)
        rate_path = rate_path.resolve()
        print(f"\nReport will use:\n  CSV:  {csv_path}\n  Rate: {rate_path}\n")
        combined, _c_issues = combine_ocean_rates_csv(csv_path)
        rate = _load_json(rate_path)
        label_src_stem = csv_path.stem
        label = args.rate_card_label.strip() or label_src_stem
    else:
        combined_path = args.update.resolve()
        rate_path = args.rate.resolve()
        if not combined_path.is_file() or not rate_path.is_file():
            print("Combined JSON or rate JSON path is missing or not a file.")
            raise SystemExit(1)
        combined = _load_json(combined_path)
        rate = _load_json(rate_path)
        label_src_stem = combined_path.stem
        label = args.rate_card_label.strip() or label_src_stem
        print(f"\nReport will use:\n  Combined: {combined_path}\n  Rate:     {rate_path}\n")

    rows_win = build_impact_rows_by_window(combined, rate, label)
    rows_cost = build_impact_rows_by_cost(combined, rate, label)

    out_name = args.output_name.strip()
    if not out_name:
        out_name = f"update_impact_{safe_json_name(label_src_stem)}.xlsx"
    if not out_name.lower().endswith(".xlsx"):
        out_name += ".xlsx"

    out_path = args.out_dir / out_name
    write_impact_workbook(out_path, rows_win, rows_cost)
    print(
        f"Wrote sheets 'By update window' ({len(rows_win)} row(s)), "
        f"'By cost' ({len(rows_cost)} row(s)) -> {out_path}"
    )


if __name__ == "__main__":
    main()
