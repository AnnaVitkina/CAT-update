"""
Convert files under input/rate (Excel) and input/update (CSV) to JSON
under processing/rate and processing/update.

Run without arguments for interactive file selection, or use --all.

Optional: ``--frames-dir DIR`` writes pandas CSV extracts per rate file
(``*_lanes.csv``, ``*_cost_columns.csv``, ``*_costs_long.csv``). Requires pandas.
Cost titles still come from openpyxl merged-cell resolution — plain ``read_excel`` is not enough.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent
INPUT_RATE = ROOT / "input" / "rate"
INPUT_UPDATE = ROOT / "input" / "update"
OUT_RATE = ROOT / "processing" / "rate"
OUT_UPDATE = ROOT / "processing" / "update"


def _json_value(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, float):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        return v
    return str(v)


def _blank_cell(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _lane_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _price_value(v):
    if _blank_cell(v):
        return None
    j = _json_value(v)
    if isinstance(j, (int, float)):
        return j
    if isinstance(j, str):
        s = j.strip()
        try:
            if "." in s:
                return float(s)
            return int(s)
        except ValueError:
            return j
    return j


def _parse_validity_block(text) -> tuple[str, str, str]:
    if text is None or (isinstance(text, str) and text.strip() == ""):
        return "", "", ""
    t = str(text).replace("\r\n", "\n").replace("\r", "\n")
    validity_parts: list[str] = []
    prolong = ""
    apply_if = ""
    for line in t.split("\n"):
        s = line.strip()
        if not s:
            continue
        low = s.lower()
        if low.startswith("validity period:"):
            validity_parts.append(s.split(":", 1)[1].strip())
        elif low.startswith("cost to prolong:"):
            prolong = s.split(":", 1)[1].strip()
        elif "applies if" in low or low.startswith("applies "):
            apply_if = s
    validity = " ".join(validity_parts) if validity_parts else ""
    return validity, prolong, apply_if


def _parse_rate_by_rule(text) -> tuple[str, str]:
    if text is None or (isinstance(text, str) and text.strip() == ""):
        return "", ""
    t = str(text).replace("\r\n", "\n").replace("\r", "\n")
    parts = [p.strip() for p in t.split("\n") if p.strip()]
    if not parts:
        return "", ""
    first = parts[0]
    rate_by = first[len("Rate by:") :].strip() if first.lower().startswith("rate by:") else first
    rule = parts[1] if len(parts) > 1 else ""
    return rate_by, rule


def _find_lane_header_row_1based(ws: Worksheet) -> int | None:
    last_row = min((ws.max_row or 250) + 1, 400)
    for r in range(1, last_row):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == "Lane #":
            return r
    return None


def iter_merged_cell_ranges(ws: Worksheet):
    try:
        yield from list(ws.merged_cells.ranges)
    except (AttributeError, TypeError, ValueError):
        pass
    try:
        for m in ws.merged_cells:
            yield m
    except (AttributeError, TypeError, ValueError):
        pass


def merged_cell_top_left_value(ws: Worksheet, row: int, col: int):
    """
    Cost titles sit in merged Currency|Price cells; only the anchor holds the value.
    Also used when openpyxl's used-range / max_column under-reports wide sheets.
    """
    for m in iter_merged_cell_ranges(ws):
        try:
            if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
                return ws.cell(row=m.min_row, column=m.min_col).value
        except (AttributeError, TypeError, ValueError):
            continue
    return ws.cell(row=row, column=col).value


def effective_max_column(ws: Worksheet) -> int:
    """
    ``worksheet.max_column`` is often too small when headers only exist inside merges or
    dimension metadata is stale — then scanning stops before RoRo / BASE_STAT_FRK columns.
    """
    m = int(ws.max_column or 0)
    try:
        dim = ws.calculate_dimension()
        if isinstance(dim, str) and ":" in dim:
            tail = dim.split(":", 1)[1]
            letters = "".join(ch for ch in tail if ch.isalpha())
            if letters:
                from openpyxl.utils import column_index_from_string

                m = max(m, column_index_from_string(letters))
    except (AttributeError, TypeError, ValueError):
        pass
    for rng in iter_merged_cell_ranges(ws):
        try:
            m = max(m, int(rng.max_col))
        except (AttributeError, TypeError, ValueError):
            continue
    return max(m, 400)


def merge_anchor_cell(ws: Worksheet, row: int, col: int) -> bool:
    """
    True if ``col`` is the **left edge** of the merge that contains ``(row, col)``.

    Requiring the **sheet** top-left ``(min_row, min_col)`` breaks real files: Transporeon
    often uses **multi-row** merged blocks for the cost header band, so the title row you scan
    is not ``min_row`` and every column was skipped — empty ``cost_names``.
    """
    for rng in iter_merged_cell_ranges(ws):
        try:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return col == rng.min_col
        except (AttributeError, TypeError, ValueError):
            continue
    return True


def _parse_rate_card_sheet_ws(sheet_name: str, ws: Worksheet) -> dict:
    """Parse Rate card using merge-aware header reads (required for Transporeon exports)."""
    h = _find_lane_header_row_1based(ws)
    if h is None or h < 5:
        raw_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        return {
            "name": sheet_name,
            "cost_names": [],
            "lanes": [],
            "parse_note": "Lane # row not found; sheet left uninterpreted.",
            "rows": [[_json_value(c) for c in row] for row in raw_rows],
        }

    cost_r = h - 4
    valid_r = h - 3
    rate_r = h - 2
    max_col = effective_max_column(ws)

    hdr_row = [ws.cell(row=h, column=c + 1).value for c in range(10)]
    shipment_count = 10
    shipment_keys: list[str] = []
    for i in range(shipment_count):
        k = hdr_row[i] if i < len(hdr_row) else None
        shipment_keys.append(str(k).strip() if k is not None else f"_column_{i}")

    cost_specs: list[dict] = []
    for c_idx in range(shipment_count, max_col):
        col_excel = c_idx + 1
        if not merge_anchor_cell(ws, cost_r, col_excel):
            continue
        name = merged_cell_top_left_value(ws, cost_r, col_excel)
        if name is None or str(name).strip() == "":
            continue
        if c_idx + 2 > max_col:
            continue
        vtext = merged_cell_top_left_value(ws, valid_r, col_excel)
        rtext = merged_cell_top_left_value(ws, rate_r, col_excel)
        validity, prolong, apply_if = _parse_validity_block(vtext)
        rate_by, rule = _parse_rate_by_rule(rtext)
        cost_specs.append(
            {
                "col": c_idx,
                "currency_col_1based": c_idx + 1,
                "cost_name": str(name).strip(),
                "validity_period": validity,
                "cost_to_prolong": prolong,
                "apply_if": apply_if,
                "rate_by": rate_by,
                "rule": rule,
            }
        )

    lanes: list[dict] = []
    for r in range(h + 1, (ws.max_row or 0) + 1):
        lane_cell = ws.cell(row=r, column=1).value
        if lane_cell is None or str(lane_cell).strip() == "":
            break

        lane_values: dict[str, str] = {}
        for i, key in enumerate(shipment_keys):
            raw = ws.cell(row=r, column=i + 1).value
            lane_values[key] = _lane_text(raw)

        costs: list[dict] = []
        for spec in cost_specs:
            c = spec["col"]
            curr = ws.cell(row=r, column=c + 1).value
            prc = ws.cell(row=r, column=c + 2).value
            if _blank_cell(curr) and _blank_cell(prc):
                continue
            costs.append(
                {
                    "Cost": spec["cost_name"],
                    "Validity period": spec["validity_period"],
                    "Cost to prolong": spec["cost_to_prolong"],
                    "Apply if": spec["apply_if"],
                    "Rate by": spec["rate_by"],
                    "Rule": spec["rule"],
                    "Currency": _lane_text(curr) if not _blank_cell(curr) else None,
                    "Price": _price_value(prc),
                }
            )

        if not costs:
            continue
        lanes.append(
            {
                "excel_row": r,
                "Lane #": lane_values.get(shipment_keys[0], ""),
                "KEY": lane_values.get("KEY", ""),
                "Rate Card": lane_values.get("Rate Card", ""),
                "Carrier": lane_values.get("Carrier", ""),
                "SERVICE": lane_values.get("SERVICE", ""),
                "Service": lane_values.get("Service", ""),
                "Valid from": lane_values.get("Valid from", ""),
                "Valid to": lane_values.get("Valid to", ""),
                "Origin Port": lane_values.get("Origin Port", ""),
                "Destination Port": lane_values.get("Destination Port", ""),
                "Costs": costs,
            }
        )

    return {
        "name": sheet_name,
        "lane_header_row": h,
        "cost_title_row": cost_r,
        "validity_header_row": valid_r,
        "rate_by_header_row": rate_r,
        "max_column": max_col,
        "shipment_column_names": shipment_keys,
        "cost_columns_parsed": len(cost_specs),
        "cost_names": [s["cost_name"] for s in cost_specs],
        "cost_specs": cost_specs,
        "lanes": lanes,
    }


def workbook_to_json(path: Path) -> dict:
    # read_only=False: merged cells are not available in read-only mode; cost titles would be missing.
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        sheets_out: dict[str, dict] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
            if sheet_name.strip().casefold() == "rate card":
                sheets_out[sheet_name] = _parse_rate_card_sheet_ws(sheet_name, ws)
            else:
                sheets_out[sheet_name] = {
                    "name": sheet_name,
                    "rows": [[_json_value(c) for c in row] for row in raw_rows],
                }
        return {
            "source_file": path.name,
            "source_path": str(path.relative_to(ROOT)),
            "sheets": sheets_out,
            "sheet_order": list(wb.sheetnames),
        }
    finally:
        wb.close()


def csv_to_json(path: Path) -> dict:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return {
            "source_file": path.name,
            "source_path": str(path.relative_to(ROOT)),
            "columns": [],
            "records": [],
        }
    header = [h.strip() if h is not None else "" for h in rows[0]]
    records = []
    for data_row in rows[1:]:
        rec = {}
        for i, col in enumerate(header):
            key = col if col else f"_column_{i}"
            val = data_row[i] if i < len(data_row) else ""
            rec[key] = val if val != "" else None
        records.append(rec)
    return {
        "source_file": path.name,
        "source_path": str(path.relative_to(ROOT)),
        "columns": header,
        "records": records,
    }


def safe_json_name(stem: str) -> str:
    s = re.sub(r'[<>:"/\\|?*]', "_", stem)
    return s.strip() or "output"


def rate_card_excel_to_dataframes(path: Path):
    """
    Build pandas tables from the Rate card (recommended workflow vs giant JSON).

    Returns ``(lanes_shipment_df, cost_columns_meta_df, costs_long_df)``.

    - **lanes_shipment_df**: one row per lane (no nested costs).
    - **cost_columns_meta_df**: one row per cost column (name, excel col index, validity, rate by).
    - **costs_long_df**: one row per lane per non-empty cost cell.

    Requires ``pandas`` (`pip install pandas`). Merged cells are still resolved via openpyxl
    in ``_parse_rate_card_sheet_ws`` — ``read_excel`` alone would not fix titles.
    """
    import pandas as pd

    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb["Rate card"]
    except KeyError:
        wb.close()
        raise ValueError("Workbook has no 'Rate card' sheet.") from None
    parsed = _parse_rate_card_sheet_ws("Rate card", ws)
    wb.close()

    lane_rows: list[dict] = []
    long_costs: list[dict] = []
    for lane in parsed.get("lanes") or []:
        base = {k: v for k, v in lane.items() if k != "Costs"}
        lane_rows.append(base)
        for co in lane.get("Costs") or []:
            long_costs.append({**base, **co})

    lanes_df = pd.DataFrame(lane_rows)
    meta_df = pd.DataFrame(parsed.get("cost_specs") or [])
    long_df = pd.DataFrame(long_costs)
    return lanes_df, meta_df, long_df


def update_csv_to_dataframe(path: Path):
    """Load Salesforce-style update CSV as a DataFrame (needs pandas)."""
    import pandas as pd

    return pd.read_csv(path, encoding="utf-8-sig", dtype=object, keep_default_na=False)


def list_files(folder: Path, patterns: tuple[str, ...]) -> list[Path]:
    if not folder.is_dir():
        return []
    out: list[Path] = []
    for pat in patterns:
        out.extend(sorted(folder.glob(pat)))
    return out


def prompt_choice(label: str, files: list[Path]) -> list[Path]:
    if not files:
        print(f"No files in {label}.")
        return []
    print(f"\n{label}:")
    for i, p in enumerate(files, start=1):
        print(f"  [{i}] {p.name}")
    raw = input(
        f"Which {label.lower()} files? (comma-separated numbers, 'a' for all, Enter to skip): "
    ).strip()
    if not raw:
        return []
    if raw.lower() == "a":
        return list(files)
    chosen: list[Path] = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            idx = int(part)
        except ValueError:
            print(f"  Ignoring invalid token: {part!r}")
            continue
        if 1 <= idx <= len(files):
            chosen.append(files[idx - 1])
        else:
            print(f"  Out of range: {idx}")
    return chosen


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert input rate/update files to JSON.")
    parser.add_argument(
        "--all",
        action="store_true",
        help="Convert every file in input/rate and input/update.",
    )
    parser.add_argument(
        "--frames-dir",
        type=Path,
        default=None,
        metavar="DIR",
        help="Also write CSV tables from pandas (lanes, cost_columns_meta, costs_long per rate file). Requires pandas.",
    )
    args = parser.parse_args()

    rate_files = list_files(INPUT_RATE, ("*.xlsx", "*.xlsm"))
    update_files = list_files(INPUT_UPDATE, ("*.csv",))

    if args.all:
        selected_rate = rate_files
        selected_update = update_files
    else:
        selected_rate = prompt_choice("input/rate (Excel)", rate_files)
        selected_update = prompt_choice("input/update (CSV)", update_files)

    OUT_RATE.mkdir(parents=True, exist_ok=True)
    OUT_UPDATE.mkdir(parents=True, exist_ok=True)

    for path in selected_rate:
        data = workbook_to_json(path)
        out_name = safe_json_name(path.stem) + ".json"
        out_path = OUT_RATE / out_name
        out_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"Wrote {out_path.relative_to(ROOT)}")

        if args.frames_dir:
            args.frames_dir.mkdir(parents=True, exist_ok=True)
            stem = safe_json_name(path.stem)
            try:
                lanes_df, meta_df, long_df = rate_card_excel_to_dataframes(path)
                lanes_df.to_csv(args.frames_dir / f"{stem}_lanes.csv", index=False)
                meta_df.to_csv(args.frames_dir / f"{stem}_cost_columns.csv", index=False)
                long_df.to_csv(args.frames_dir / f"{stem}_costs_long.csv", index=False)
                print(
                    f"Wrote {args.frames_dir.resolve() / (stem + '_*.csv')} (pandas tables)"
                )
            except ImportError as e:
                print(f"  Skipping --frames-dir: {e}")

    for path in selected_update:
        data = csv_to_json(path)
        out_name = safe_json_name(path.stem) + ".json"
        out_path = OUT_UPDATE / out_name
        out_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"Wrote {out_path.relative_to(ROOT)}")

    if not selected_rate and not selected_update:
        print("Nothing selected; no output written.")


if __name__ == "__main__":
    main()
