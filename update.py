"""
Ocean rates **update** CSV + **rate** JSON (from Excel) → combined records → merge → Excel result.

Reference workflow (container CY-CY, e.g. ``Ocean Rates 20260306 23-24`` ×
``Advanced Export - RA20250626007 v.23 - Sea``):

1. **Lane identity** — Strip trailing equipment from ``KEY`` for routing (e.g.
   ``CMDU-OC_CNTR_CY-CY_BU-MXATM-ZADUR-42P1B`` → route ``…-MXATM-ZADUR``, equipment ``42P1B``).
   Match lanes on Carrier, Service (from ``SERVICE__C``, e.g. CY-CY), Origin/Destination ports,
   and validity overlap where splits apply.

2. **Costs from CSV** — ``42xxCNTR_*`` columns → display names
   ``Base Rate (42xx)``, ``Destination Terminal Handling Fee (42xx)``, …;
   BAF dated bands ``BAF Fee (42xx dd.mm.yyyy–dd.mm.yyyy)`` from effective/expiration dates.

3. **No matching lane** — Append a new lane row; insert missing cost columns (**2** or **3** wide
   per MIN) **right-to-left** so existing columns left of each insert stay fixed until shifted as whole columns.

4. **BAF tiers** — If an existing rate column’s validity **fully covers** the CSV window (e.g.
   ``BAF Fee (42P1 15.01.2026-31.05.2026)`` covers ``23.02.2026–31.05.2026``), write prices into
   that tier column (not a new dated title).

5. **Existing lane, new equipment** — e.g. lane matches route but had no ``42UTH`` costs: add new cost
   columns and rows without unnecessary validity trimming when the spec says “only add tiers.”

6. **Lane-head validity split (*DFT* w/m)** — If the CSV window **starts on the same day** as the lane
   but **ends earlier** than the lane (e.g. lane ``01.11.2025–31.10.2026``, update ``01.11.2025–06.03.2026``),
   the merge keeps the **same Lane #** through the CSV end with merged costs and **appends** one lane
   duplicating pre-split costs for the remainder (second CSV row then updates that continuation lane).

7. **DFTBASE / DFTROLL Minimum** — Non-zero ``*_MIN`` becomes ``Minimum`` on the JSON cost row.
   In Excel, costs with MIN use **three** columns: **Currency** | **Flat** (MIN value; header shows ``MIN`` + ``Flat``)
   | **p/unit**. Zero MIN keeps the two-column Currency | p/unit layout (no Flat column).

Additional code paths support other RoRo / tier-edge cases; container CY-CY and *DFT* breakbulk above
are the documented reference flows. *DFT* BAF / EU ETS ``since``→dated column migration (2026 window
splits) runs only when the CSV for this merge includes *DFT* BAF / EU ETS fee values; see
``MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES``.

**Inputs:** *Rate* — ``processing/rate/*.json``; *Update* — ``input/update/*.csv``.

Pipeline: (1) combine CSV rows (same KEY prefix); (2) merge JSON lanes; (3) Excel from
template with green (changed) / red (new structure) fills.

Set ``MERGE_FULL_TRACE = False`` for shorter ``*_report.txt``. Use ``--debug`` for stderr logging detail.
"""
from __future__ import annotations

import argparse
import copy
import logging
from collections import defaultdict
import json
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.cell_range import CellRange

from transform_inputs import (
    ROOT,
    csv_to_json,
    effective_max_column,
    merge_anchor_cell,
    merged_cell_top_left_value,
    safe_json_name,
)

INPUT_UPDATE = ROOT / "input" / "update"
INPUT_RATE = ROOT / "input" / "rate"
OUT_COMBINED_DIR = ROOT / "processing" / "update_to_perform"
OUT_RESULT_DIR = ROOT / "processing" / "result"

# Verbose per-record lines in ``processing/result/*_report.txt`` (merge trace).
MERGE_FULL_TRACE = True

# *DFT* BAF / EU ETS: ``since`` headers → two dated windows (deep-copy row; same Currency/Price).
# By default this runs only when the **current CSV** includes *DFT* BAF / EU ETS fee values
# (``*DFTBAF_*`` / ``*DFTEU_ETS_*``), so container-only updates do not restructure breakbulk
# fee columns. Set ``MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES = False`` to always
# migrate every lane that still has a ``since`` row (legacy card-wide behavior).
MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES = True
BAF_SINCE_2026_COST = "BAF Fee (since 15.01.2026)"
BAF_SPLIT_WINDOWS_2026: tuple[tuple[str, date, date], ...] = (
    ("BAF Fee (15.01.2026 - 14.04.2026)", date(2026, 1, 15), date(2026, 4, 14)),
    ("BAF Fee (15.04.2026 - 31.10.2026)", date(2026, 4, 15), date(2026, 10, 31)),
)
EU_ETS_SINCE_2026_COST = "EU ETS Fee (since 15.01.2026)"
EU_ETS_SPLIT_WINDOWS_2026: tuple[tuple[str, date, date], ...] = (
    ("EU ETS Fee (01.01.2026 - 31.03.2026)", date(2026, 1, 1), date(2026, 3, 31)),
    ("EU ETS Fee (01.04.2026 - 31.10.2026)", date(2026, 4, 1), date(2026, 10, 31)),
)

# When ``True``, CSV rows are merged into the rate JSON even if they pass the redundant
# tier check (lane already has matching BAF/EU ETS price+currency for the window). Default
# ``False`` keeps skipping duplicate-tier updates; set ``True`` to always refresh lanes /
# bump Excel when columns already exist.
APPLY_CSV_UPDATES_EVEN_WHEN_REDUNDANT = False

log = logging.getLogger("update_pipeline")
_update_log_handlers = False


def configure_update_logging(debug: bool) -> None:
    """Configure stderr logging for this pipeline (WARNING unless ``debug``)."""
    global _update_log_handlers
    log.setLevel(logging.DEBUG if debug else logging.WARNING)
    if not _update_log_handlers:
        h = logging.StreamHandler()
        h.setFormatter(logging.Formatter("[%(levelname)s] update: %(message)s"))
        log.addHandler(h)
        log.propagate = False
        _update_log_handlers = True


def parse_agreement_and_version_from_rate(prior: dict) -> tuple[str | None, int | None]:
    rows = prior.get("sheets", {}).get("General info", {}).get("rows", [])
    agreement: str | None = None
    ver_num: int | None = None
    for row in rows:
        if len(row) < 2:
            continue
        k, v = row[0], row[1]
        if k == "Agreement number" and v:
            agreement = str(v).strip()
        if k == "Version" and v:
            m = re.search(r"(?i)v\.(\d+)", str(v))
            if m:
                ver_num = int(m.group(1))
    return agreement, ver_num


def bump_version_string_in_general_info(rate: dict) -> None:
    """Increment ``v.N`` in General info → Excel export shows bumped workbook version."""
    rows = rate.get("sheets", {}).get("General info", {}).get("rows", [])
    for row in rows:
        if len(row) < 2 or row[0] != "Version" or not row[1]:
            continue

        def up(m):
            return f"v.{int(m.group(1)) + 1}"

        new_val = re.sub(r"(?i)v\.(\d+)", up, str(row[1]), count=1)
        row[1] = new_val
        log.debug(
            "General info Version bumped for merged JSON output (Excel reflects next agreement version): %r",
            new_val,
        )
        break


def result_stem_for_merge_output(
    rate_json_path: Path, template_path: Path, prior: dict
) -> str:
    """
    Stem for ``processing/result/*.json`` and ``*.xlsx``.

    Prefer bumping ``v.N`` in the **selected rate JSON** filename (what you pick as the
    merge baseline). If it has no ``v.N``, use the template workbook name, then
    ``General info`` Version / agreement fallbacks.
    """
    agreement, ver = parse_agreement_and_version_from_rate(prior)

    def repl(m):
        return f"v.{int(m.group(1)) + 1}"

    for stem in (rate_json_path.stem, template_path.stem):
        new_stem, n = re.subn(r"(?i)v\.(\d+)", repl, stem, count=1)
        if n:
            return safe_json_name(new_stem)
    if agreement and ver is not None:
        return safe_json_name(f"{agreement} v.{ver + 1}")
    return safe_json_name(f"{rate_json_path.stem} updated")


META_COMPARE_COLS = [
    "CARRIER",
    "CARRIERNAME",
    "SERVICE__C",
    "COMMODITY__C",
    "RATE_EFFECTIVE_DATE__C",
    "RATE_EXPIRATION_DATE__C",
    "SERVICE_GRADE_NUMBER__C",
    "SERVICE_GRADE_DEFINITION",
    "ORIGIN_REGION__C",
    "ORIGIN_LOCATION_NAME__C",
    "ORIGIN_COUNTRY__C",
    "ORIGIN_CITY__C",
    "ORIGIN_STATE__C",
    "ORIGIN_ZONE__C",
    "DESTINATION_REGION__C",
    "DESTINATION_LOCATION_NAME__C",
    "DESTINATION_COUNTRY__C",
    "DESTINATION_CITY__C",
    "DESTINATION_STATE__C",
    "DESTINATION_ZONE__C",
]

DATE_EXCEPTION_COLS = frozenset({"RATE_EFFECTIVE_DATE__C", "RATE_EXPIRATION_DATE__C"})

# Metadata that must match across CSV rows merged into one combined record (dates excluded).
_META_MERGE_LOCK_COLS = frozenset(
    c for c in META_COMPARE_COLS if c not in DATE_EXCEPTION_COLS
)

# Breakbulk *DFT* BAF / EU ETS: when several CSV windows are merged for the same route, each
# window is stored as one element of ``DFT_BAF_EU_ETS_WINDOWS`` (valid JSON; no repeated keys).
DFT_BAF_EU_ETS_WINDOWS_KEY = "DFT_BAF_EU_ETS_WINDOWS"
DFT_BAF_EU_ETS_FLAT_COLS = frozenset(
    {
        "*DFTBAF_Currency",
        "*DFTBAF_MIN",
        "*DFTBAF_rate (per w/m)",
        "*DFTEU_ETS_Currency",
        "*DFTEU_ETS_MIN",
        "*DFTEU_ETS_rate (per w/m)",
    }
)

NON_BAF_KIND_LABEL = {
    "BASE": "Base Rate ({eqp})",
    "DTHC": "Destination Terminal Handling Fee ({eqp})",
    "OTHC": "Origin Terminal Handling Fee ({eqp})",
}

# Lower = further left when grouping cost columns by equipment (see ``_sort_key_cost_column_name``).
EQ_SORT_ORDER = {
    "45G0": 0,
    "45SOC": 1,
    "42P1B": 10,
    "42P1H": 11,
    "42P1": 12,
    "42UTH": 20,
    "42G0": 13,
    "22G0": 30,
    "22P1": 31,
    "22P1B": 32,
    "22UTH": 33,
}

# *DFT* breakbulk: CSV windows **longer** than this (days) are treated as a single negotiated
# rate on ``BAF Fee (since …)`` / ``EU ETS Fee (since …)`` (in-place price). Shorter windows
# carve a dated band + trailing ``since`` (Ocean Rates 20260327).
DFT_SINCE_INPLACE_MIN_SPAN_DAYS = 100

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _csv_port_matches_lane_field(csv_loc: str, lane_port_field: str) -> bool:
    """
    CSV usually has one UN/LOCODE (e.g. ``CNSHA``); rate card ``Origin Port`` / ``Destination Port``
    may list aliases ``CNSHA/CNSHG/CNSGH``.
    """
    c = _norm(csv_loc).upper()
    lane = _norm(lane_port_field).upper()
    if not c or not lane:
        return False
    if lane == c:
        return True
    if "/" in lane:
        return any(p.strip() == c for p in lane.split("/"))
    if "/" in c:
        return any(p.strip() == lane for p in c.split("/"))
    return False


def _parse_us_date(s) -> date | None:
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_us_mdy(d: date) -> str:
    """US slash dates like source CSVs (e.g. ``1/15/2026``)."""
    return f"{d.month}/{d.day}/{d.year}"


def _fmt_dmy(d: date) -> str:
    return f"{d.day:02d}.{d.month:02d}.{d.year}"


def _parse_dmy_text(s: str) -> date | None:
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s.strip())
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_validity_range(period: str) -> tuple[date | None, date | None]:
    if not period:
        return None, None
    period = period.strip()
    m = re.search(
        r"from\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+to\s+(\d{1,2}\.\d{1,2}\.\d{4})",
        period,
        re.I,
    )
    if m:
        return _parse_dmy_text(m.group(1)), _parse_dmy_text(m.group(2))
    m_since = re.search(r"since\s+(\d{1,2}\.\d{1,2}\.\d{4})", period, re.I)
    if m_since:
        return _parse_dmy_text(m_since.group(1)), None
    m2 = re.search(r"to\s+(\d{1,2}\.\d{1,2}\.\d{4})", period, re.I)
    if m2:
        return None, _parse_dmy_text(m2.group(1))
    m3 = re.match(
        r"^from\s+(\d{1,2}\.\d{1,2}\.\d{4})\s*$",
        period.strip(),
        re.I,
    )
    if m3:
        return _parse_dmy_text(m3.group(1)), None
    return None, None


def _tier_validity_bounds(t: dict) -> tuple[date | None, date | None]:
    """
    Calendar bounds for a cost row — **only** from ``Validity period`` (``from … to …``,
    ``since …``, open-ended ``from …``, etc.). The ``Cost`` column title is **not** parsed
    for dates (titles remain display / Excel identity only).
    """
    return _parse_validity_range(t.get("Validity period") or "")


def _dft_breakbulk_since_style_validity(t: dict) -> bool:
    """
    Open-ended *DFT* BAF / EU ETS tier: detected from **Validity period** — text contains
    ``since dd.mm.yyyy``, or parses to a start date with no end (open band).
    """
    p = (t.get("Validity period") or "").strip()
    if not p:
        return False
    if re.search(r"\bsince\s+\d{1,2}\.\d{1,2}\.\d{4}", p, re.I):
        return True
    d1, d2 = _parse_validity_range(p)
    return d1 is not None and d2 is None


def _dft_semantic_validity_bounds(t: dict, eqp: str) -> tuple[date | None, date | None]:
    """Same as :func:`_tier_validity_bounds`; ``eqp`` kept for call-site compatibility."""
    return _tier_validity_bounds(t)


def _fee_tier_semantic_open_since_bounds(
    t: dict,
    eqp: str,
    fee_title: str,
    match_equipment,
) -> tuple[date | None, date | None]:
    """
    Semantic bounds for tier matching — **Validity period** only (see :func:`_tier_validity_bounds`).
    ``fee_title`` / ``match_equipment`` retained for API compatibility with callers.
    """
    return _tier_validity_bounds(t)


def _tier_validity_overlaps_window(t: dict, w_lo: date, w_hi: date) -> bool:
    """
    True if the tier's validity window intersects ``[w_lo, w_hi]``.
    Handles open-ended ``since`` / ``to`` rows via ``_tier_validity_bounds``.
    """
    d1, d2 = _tier_validity_bounds(t)
    if d1 is not None and d2 is not None:
        return _date_range_intersection(w_lo, w_hi, d1, d2) is not None
    if d1 is None and d2 is not None:
        return w_lo <= d2
    if d1 is not None and d2 is None:
        return d1 <= w_hi
    return True


def _filter_container_baf_block_to_csv_window(
    block: list[dict], eqp: str, eff: date, exp: date
) -> list[dict]:
    """
    Drop container-style BAF tiers whose validity does **not** overlap the CSV
    ``[eff, exp]``. Donor templates often carry historical tiers (710, 809, …);
    the update only supplies one rate for the current window — stale tiers must
    not remain on the merged lane.
    """
    if eqp == "*DFT":
        return block
    out: list[dict] = []
    for c in block:
        nm = c.get("Cost") or ""
        if not baf_matches_equipment(nm, eqp):
            out.append(c)
            continue
        if not _is_container_style_baf(nm):
            out.append(c)
            continue
        if _tier_validity_overlaps_window(c, eff, exp):
            out.append(c)
    return out


def _price_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v) if isinstance(v, float) and v.is_integer() else v
    s = str(v).strip()
    try:
        return int(s) if "." not in s else float(s)
    except ValueError:
        return s


def _numeric_fee_equal(a, b) -> bool:
    """Compare CSV / lane fee amounts without float noise."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return float(_price_num(a)) == float(_price_num(b))
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


def lane_prefix_from_key(key: str) -> str:
    """Routing prefix for merge/combine: strip trailing equipment (container code or ``*DFT``)."""
    if not key:
        return ""
    suf = equipment_suffix_from_key(key)
    if suf:
        return key[: -(len(suf) + 1)]
    parts = key.rsplit("-", 1)
    if len(parts) == 2 and re.match(r"^[A-Za-z0-9]+$", parts[1]):
        return parts[0]
    return key


def equipment_suffix_from_key(key: str) -> str | None:
    """
    Trailing **equipment** token from an update KEY, e.g. ``…-42UTH`` → ``42UTH``,
    ``…-*DFT`` → ``*DFT``.

    The last ``-`` segment is **not** equipment when it is a port code (e.g. ``…-SGSIN-ZADUR`` —
    the final ``ZADUR`` is destination, not a container type). Only ``*DFT`` or
    ``[0-9]{2}…`` container-style codes are treated as equipment.
    """
    if not key:
        return None
    parts = key.rsplit("-", 1)
    if len(parts) != 2:
        return None
    suf = parts[1].strip()
    if not re.match(r"^[A-Za-z0-9*]+$", suf):
        return None
    if suf == "*DFT":
        return suf
    if re.match(r"^[0-9]{2}[A-Za-z0-9]+$", suf):
        return suf
    return None


def _lane_mentions_equipment(lane: dict, eq: str) -> bool:
    """True if any ``Cost`` label references ``eq`` as container equipment (in parentheses)."""
    if not eq or eq == "*DFT":
        return False
    for c in lane.get("Costs") or []:
        nm = str(c.get("Cost") or "")
        if f"({eq}" in nm or f"({eq} " in nm:
            return True
    return False


def _route_has_equipment_on_any_lane(route_lanes: list[dict], eq: str) -> bool:
    return any(_lane_mentions_equipment(ln, eq) for ln in route_lanes)


def eq_paren(eq: str) -> str:
    """
    Equipment code placed in ``Cost`` labels (``Base Rate ({eqp})``, ``BAF Fee ({eqp})``, …).
    Must match the CSV column prefix (e.g. ``42UTHCNTR_*`` → ``42UTH``), not a shortened alias.
    """
    return eq


def eq_sort_key(eq: str) -> tuple[int, str]:
    return (EQ_SORT_ORDER.get(eq, 99), eq)


def service_cy_cy(service__c: str) -> str:
    """
    Normalize ``SERVICE__C`` for lane matching (must equal lane ``Service``).

    - ``OC_CNTR_CY-CY_BU`` → ``CY-CY`` (segment between ``OC_CNTR_`` and ``_BU``).
    - Else first ``CY-…`` substring.
    - Else ``OC_<CODE>`` → ``RRBB`` etc. (e.g. ``OC_RRBB``).
    """
    s = service__c or ""
    m0 = re.search(
        r"OC_CNTR_((?:CY-[A-Z0-9]+(?:-[A-Z0-9]+)*))_BU",
        s,
        re.I,
    )
    if m0:
        return m0.group(1).upper()
    m = re.search(r"(CY-[A-Z0-9]+(?:-[A-Z0-9]+)*)", s, re.I)
    if m:
        return m.group(1).upper()
    m2 = re.search(r"\bOC_([A-Z0-9]+)\b", s, re.I)
    if m2:
        return m2.group(1).upper()
    return ""


def equipment_code_before_cntr(col: str) -> str | None:
    """
    Equipment = everything before the first ``CNTR`` in the column name
    (e.g. ``42P1BCNTR_DTHC_MIN`` → ``42P1B``, ``42UTHCNTR_…_rate`` → ``42UTH``).
    """
    if not col:
        return None
    idx = col.upper().find("CNTR")
    if idx <= 0:
        return None
    eq = col[:idx].strip()
    return eq or None


def rate_column_eq_kind(col: str) -> tuple[str, str] | None:
    """
    ``{equipment}CNTR_{BAF|BASE|DTHC|OTHC}_rate (per container)``.

    ``equipment`` is the substring **before** ``CNTR`` (see ``equipment_code_before_cntr``),
    not a separate regex on digits—so ``42P1B`` vs ``42P1`` disambiguate correctly.
    """
    eq = equipment_code_before_cntr(col)
    if not eq:
        return None
    idx = col.upper().find("CNTR")
    tail = col[idx:]
    m = re.match(r"CNTR_(BAF|BASE|DTHC|OTHC)_rate \(per container\)$", tail, re.I)
    if not m:
        return None
    return eq, m.group(1).upper()


def dft_rate_column_kind(col: str) -> tuple[str, str] | None:
    """Breakbulk *DFT columns: ``*DFTBAF_*`` / ``*DFTEU_ETS_*`` → (``*DFT``, ``BAF``|``EU_ETS``)."""
    m = re.match(r"^\*DFT(BAF|EU_ETS)_rate \(per w/m\)$", col)
    if not m:
        return None
    fk = "EU_ETS" if m.group(1) == "EU_ETS" else "BAF"
    return "*DFT", fk


def dft_breakbulk_base_rate_cost_label(col: str) -> str | None:
    """
    ``*DFTBASE_STAT_FRK_rate (per w/m)`` → ``Base Rate (BASE_STAT_FRK)``;
    ``*DFTROLL_TR_NONFRK_rate (per w/m)`` → ``Base Rate (DFTROLL_TR_NONFRK)``.
    """
    m = re.match(r"^\*DFT(BASE|ROLL)_([A-Za-z0-9_]+)_rate \(per w/m\)$", col)
    if not m:
        return None
    kind, tier = m.group(1), m.group(2)
    if kind == "BASE":
        return f"Base Rate (BASE_{tier})"
    return f"Base Rate (DFTROLL_{tier})"


def _dft_breakbulk_min_from_rec(rec: dict, rate_col: str) -> object | None:
    v = rec.get(rate_col.replace("_rate (per w/m)", "_MIN"))
    if v in (None, ""):
        return None
    n = _price_num(v)
    if isinstance(n, (int, float)) and n == 0:
        return None
    return n


def build_dft_breakbulk_wm_base_cost_rows(
    rec: dict,
    all_lanes: list[dict],
    issues: list[str],
    *,
    validity_eff_exp: tuple[date | None, date | None] | None = None,
) -> list[dict]:
    """
    One cost row per ``*DFTBASE_*`` / ``*DFTROLL_*`` rate column with a value.
    Optional ``Minimum`` from ``*_MIN`` when non-zero (shown in Excel header text).
    """
    eff_v = exp_v = None
    if validity_eff_exp:
        eff_v, exp_v = validity_eff_exp
    out: list[dict] = []
    for col in rec:
        label = dft_breakbulk_base_rate_cost_label(col)
        if not label or rec.get(col) in (None, ""):
            continue
        ccy = rec.get(col.replace("_rate (per w/m)", "_Currency"))
        pr = rec[col]
        tmpl = find_template_cost(all_lanes, label)
        if tmpl:
            tmpl = copy.deepcopy(tmpl)
        else:
            tmpl = _missing_template_non_baf_cost(label)
            issues.append(
                f"  no rate-card pattern for {label!r} — "
                f"``Rate by``/``Rule`` left empty; mark Excel Rate-by cell red."
            )
        tmpl["Currency"] = ccy
        tmpl["Price"] = _price_num(pr)
        mnv = _dft_breakbulk_min_from_rec(rec, col)
        if mnv is not None:
            tmpl["Minimum"] = mnv
        if eff_v and exp_v:
            tmpl["Validity period"] = f"from {_fmt_dmy(eff_v)} to {_fmt_dmy(exp_v)}"
        out.append(tmpl)
    return out


def equipment_in_update(rec: dict) -> list[str]:
    found: set[str] = set()
    for col in rec:
        rk = rate_column_eq_kind(col)
        if rk and rec.get(col) not in (None, ""):
            found.add(rk[0])
        dk = dft_rate_column_kind(col)
        if dk and rec.get(col) not in (None, ""):
            found.add(dk[0])
        if dft_breakbulk_base_rate_cost_label(col) and rec.get(col) not in (None, ""):
            found.add("*DFT")
    return sorted(found, key=eq_sort_key)


def container_equipment_from_cntr_columns(rec: dict) -> list[str]:
    """Container codes with at least one non-empty ``*CNTR_*_rate`` column (excludes *DFT)."""
    found: set[str] = set()
    for col in rec:
        rk = rate_column_eq_kind(col)
        if rk and rec.get(col) not in (None, ""):
            eq, _ = rk
            if eq != "*DFT":
                found.add(eq)
    return sorted(found, key=eq_sort_key)


def merge_routing_container_equipment(rec: dict) -> str | None:
    """
    Equipment token for merge branching when CSV rows are combined without a ``KEY`` suffix:
    use trailing ``…-42UTH`` if present, else the sole ``*CNTR_*`` equipment code.
    """
    sk = equipment_suffix_from_key(_norm(rec.get("KEY")))
    if sk and sk != "*DFT":
        return sk
    ceqs = container_equipment_from_cntr_columns(rec)
    if len(ceqs) == 1:
        return ceqs[0]
    return None


def _is_container_style_baf(cost_name: str) -> bool:
    """``BAF Fee (42P1 …)`` etc.; excludes breakbulk (*DFT / since / date-only titles)."""
    if not cost_name.startswith("BAF Fee ("):
        return False
    rest = cost_name[len("BAF Fee (") :]
    return bool(re.match(r"^[0-9]{2}[A-Za-z]", rest))


def baf_matches_equipment(cost_name: str, eqp: str) -> bool:
    if not cost_name.startswith("BAF Fee"):
        return False
    if eqp == "*DFT":
        # Ignore ``*DFT`` as equipment: one chain for all non-container BAF (since, *DFT-dated, …).
        return not _is_container_style_baf(cost_name)
    rest = cost_name[len("BAF Fee (") :]
    return rest.startswith(eqp + ")") or rest.startswith(eqp + " ")


def _is_container_style_eu_ets(cost_name: str) -> bool:
    if not cost_name.startswith("EU ETS Fee ("):
        return False
    rest = cost_name[len("EU ETS Fee (") :]
    return bool(re.match(r"^[0-9]{2}[A-Za-z]", rest))


def eu_ets_matches_equipment(cost_name: str, eqp: str) -> bool:
    if not cost_name.startswith("EU ETS Fee"):
        return False
    if eqp == "*DFT":
        return not _is_container_style_eu_ets(cost_name)
    rest = cost_name[len("EU ETS Fee (") :]
    return rest.startswith(eqp + ")") or rest.startswith(eqp + " ")


def _is_dft_breakbulk_wm_base_cost_name(cost_name: str) -> bool:
    """
    True for *DFT* w/m base rows from ``*DFTBASE_*`` / ``*DFTROLL_*`` CSV columns —
    ``Base Rate (BASE_STAT_FRK)``, ``Base Rate (DFTROLL_TR_NONFRK)``, etc.
    Used when stripping old rows before merging CSV ``new_costs`` for KEY suffix ``*DFT``.
    """
    n = (cost_name or "").strip()
    return n.startswith("Base Rate (BASE_") or n.startswith("Base Rate (DFTROLL_")


def cost_belongs_equipment(cost_name: str, eqp: str) -> bool:
    if eqp == "*DFT":
        if cost_name.startswith("BAF Fee") and not _is_container_style_baf(cost_name):
            return True
        if cost_name.startswith("EU ETS Fee") and not _is_container_style_eu_ets(
            cost_name
        ):
            return True
        if _is_dft_breakbulk_wm_base_cost_name(cost_name):
            return True
        return False
    if f"({eqp})" in cost_name or f"({eqp} " in cost_name:
        return True
    return False


def list_update_csvs() -> list[Path]:
    if not INPUT_UPDATE.is_dir():
        return []
    return sorted(
        p for p in INPUT_UPDATE.glob("*.csv") if not p.name.startswith("~$")
    )


def list_rate_jsons() -> list[Path]:
    d = ROOT / "processing" / "rate"
    if not d.is_dir():
        return []
    return sorted(p for p in d.glob("*.json") if p.is_file())


def list_rate_templates() -> list[Path]:
    if not INPUT_RATE.is_dir():
        return []
    return sorted(
        p
        for p in INPUT_RATE.glob("*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    )


def _merge_group_key(rec: dict) -> tuple:
    """
    Merge key = lane prefix only (``KEY`` with trailing equipment stripped).

    Rows for the same route that share this prefix are folded into one combined record.
    Validity handling depends on :func:`combine_ocean_rates_csv` (window array vs union dates).
    """
    pfx = lane_prefix_from_key(_norm(rec.get("KEY")))
    return (pfx,)


def _row_only_dft_baf_eu_ets_rates_no_other_cost_columns(row: dict) -> bool:
    """
    True if this CSV row has no non-empty **other** rate columns (container ``*CNTR_*``,
    ``*DFTBASE_*`` / ``*DFTROLL_*`` w/m base, etc.). Non-empty *DFT* BAF / EU ETS columns are
    allowed; metadata and validity dates are ignored by this check.
    """
    for col, val in row.items():
        if val in (None, ""):
            continue
        if col in _META_MERGE_LOCK_COLS or col in DATE_EXCEPTION_COLS or col == "KEY":
            continue
        if col in ("source_keys", DFT_BAF_EU_ETS_WINDOWS_KEY):
            continue
        if col in DFT_BAF_EU_ETS_FLAT_COLS:
            continue
        if rate_column_eq_kind(col) or dft_breakbulk_base_rate_cost_label(col):
            return False
    return True


def _dft_fee_window_dict(row: dict) -> dict:
    """One *DFT* BAF / EU ETS window: US dates + the six w/m fee columns from a CSV row."""
    w: dict = {
        "RATE_EFFECTIVE_DATE__C": row.get("RATE_EFFECTIVE_DATE__C"),
        "RATE_EXPIRATION_DATE__C": row.get("RATE_EXPIRATION_DATE__C"),
    }
    for c in DFT_BAF_EU_ETS_FLAT_COLS:
        w[c] = row.get(c)
    return w


def expand_merged_records_for_dft_fee_windows(
    records: list[dict],
) -> list[dict]:
    """
    Expand combined records that carry ``DFT_BAF_EU_ETS_WINDOWS`` into flat rows (debug/export).

    **Merge does not use this** — :func:`merge_rate_with_updates` reads windows from each
    combined record via :func:`flat_records_from_combined_merge_rec` so new routes batch
    correctly into one lane.
    """
    out: list[dict] = []
    for rec in records:
        wins = rec.get(DFT_BAF_EU_ETS_WINDOWS_KEY)
        if not isinstance(wins, list) or not wins:
            out.append(rec)
            continue
        base = {k: v for k, v in rec.items() if k != DFT_BAF_EU_ETS_WINDOWS_KEY}
        for w in wins:
            out.append({**base, **w})
    return out


def flat_records_from_combined_merge_rec(rec: dict) -> list[dict]:
    """
    Flat CSV-like rows for one combined record: either ``[rec]`` or one dict per
    ``DFT_BAF_EU_ETS_WINDOWS`` entry (dates + *DFTBAF* / *DFTEU_ETS* columns merged onto metadata).
    """
    wins = rec.get(DFT_BAF_EU_ETS_WINDOWS_KEY)
    if isinstance(wins, list) and wins:
        base = {k: v for k, v in rec.items() if k != DFT_BAF_EU_ETS_WINDOWS_KEY}
        return [{**base, **w} for w in wins]
    return [rec]


def sort_key_combined_or_merge_record(rec: dict) -> tuple[date, date]:
    """Chronological sort key for combined JSON rows and merge input order."""
    e = _parse_us_date(str(rec.get("RATE_EFFECTIVE_DATE__C") or "").strip())
    x = _parse_us_date(str(rec.get("RATE_EXPIRATION_DATE__C") or "").strip())
    if e or x:
        return (e or date.min, x or date.min)
    wins = rec.get(DFT_BAF_EU_ETS_WINDOWS_KEY)
    if isinstance(wins, list) and wins:
        w0 = wins[0]
        return (
            _parse_us_date(str(w0.get("RATE_EFFECTIVE_DATE__C") or "").strip())
            or date.min,
            _parse_us_date(str(w0.get("RATE_EXPIRATION_DATE__C") or "").strip())
            or date.min,
        )
    return (date.min, date.min)


def combine_ocean_rates_csv(csv_path: Path) -> tuple[dict, list[str]]:
    """
    Merge CSV rows that share the same lane KEY prefix (equipment stripped from ``KEY``).

    One combined record per route. If several CSV rows share the same route and **every** row
    only carries *DFT* BAF / EU ETS rate columns (no container or *DFTBASE*/*DFTROLL* rates),
    those fees are stored in ``DFT_BAF_EU_ETS_WINDOWS`` (one object per window). The parent
    object then has **no** top-level ``RATE_EFFECTIVE_DATE__C`` / ``RATE_EXPIRATION_DATE__C``
    (validity lives only inside each window). If any row includes other cost columns, multi-row
    merge stays flat with last-wins and union top-level dates instead.
    """
    log.debug("combine_ocean_rates_csv: reading %s", csv_path)
    raw = csv_to_json(csv_path)
    records = raw["records"]
    log.debug("combine: %s raw CSV row(s) after csv_to_json.", len(records))
    issues: list[str] = []
    groups: dict[tuple, list[dict]] = {}
    for r in records:
        gk = _merge_group_key(r)
        groups.setdefault(gk, []).append(r)
    log.debug(
        "combine: folded into %s merge group(s) (same lane prefix).",
        len(groups),
    )

    merged_records: list[dict] = []
    for gk in sorted(groups.keys()):
        pfx = gk[0]
        rows = sorted(
            groups[gk],
            key=lambda x: (
                _parse_us_date(str(x.get("RATE_EFFECTIVE_DATE__C") or "").strip())
                or date.min,
                _parse_us_date(str(x.get("RATE_EXPIRATION_DATE__C") or "").strip())
                or date.min,
                _norm(x.get("KEY")),
            ),
        )
        base = rows[0]
        for r in rows[1:]:
            for col in _META_MERGE_LOCK_COLS:
                if _norm(base.get(col)) != _norm(r.get(col)):
                    issues.append(
                        f"{pfx}: metadata mismatch column {col!r}: "
                        f"{base.get('KEY')} ({base.get(col)!r}) vs {r.get('KEY')} ({r.get(col)!r})"
                    )

        merged: dict = {**base}
        merged["KEY"] = pfx
        merged["source_keys"] = list(
            dict.fromkeys(_norm(r.get("KEY")) for r in rows)
        )
        multi_window = len(rows) > 1
        use_dft_windows = multi_window and all(
            _row_only_dft_baf_eu_ets_rates_no_other_cost_columns(r) for r in rows
        )
        if use_dft_windows:
            merged[DFT_BAF_EU_ETS_WINDOWS_KEY] = [_dft_fee_window_dict(r) for r in rows]
            for col in DFT_BAF_EU_ETS_FLAT_COLS:
                merged[col] = None
            merged.pop("RATE_EFFECTIVE_DATE__C", None)
            merged.pop("RATE_EXPIRATION_DATE__C", None)
        for r in rows:
            for k, v in r.items():
                if k == "KEY":
                    continue
                if k in _META_MERGE_LOCK_COLS:
                    continue
                if k in DATE_EXCEPTION_COLS:
                    continue
                if use_dft_windows and k in DFT_BAF_EU_ETS_FLAT_COLS:
                    continue
                if v not in (None, ""):
                    merged[k] = v
        if len(rows) > 1 and not use_dft_windows:
            effs: list[date] = []
            exps: list[date] = []
            for r in rows:
                e = _parse_us_date(str(r.get("RATE_EFFECTIVE_DATE__C") or "").strip())
                x = _parse_us_date(str(r.get("RATE_EXPIRATION_DATE__C") or "").strip())
                if e is not None:
                    effs.append(e)
                if x is not None:
                    exps.append(x)
            if effs:
                merged["RATE_EFFECTIVE_DATE__C"] = _fmt_us_mdy(min(effs))
            if exps:
                merged["RATE_EXPIRATION_DATE__C"] = _fmt_us_mdy(max(exps))
        merged_records.append(merged)

    # Chronological order for consumers reading combined JSON (windows: first segment dates).
    merged_records.sort(key=sort_key_combined_or_merge_record)

    out = {
        "source_file": raw["source_file"],
        "source_path": raw["source_path"],
        "columns": list(raw["columns"]) + ["source_keys", DFT_BAF_EU_ETS_WINDOWS_KEY],
        "records": merged_records,
    }
    log.debug(
        "combine: output %s merged record(s); %s issue line(s) (metadata conflicts etc.).",
        len(merged_records),
        len(issues),
    )
    return out, issues


def find_donor_lane(lanes: list[dict], carrier: str | None) -> dict:
    if not lanes:
        return {"Lane #": "", "Costs": []}

    cand = [l for l in lanes if not carrier or _norm(l.get("Carrier")) == _norm(carrier)]
    if not cand:
        cand = lanes

    def score(ln: dict) -> int:
        names = {c.get("Cost") for c in ln.get("Costs", [])}
        return sum(1 for n in names if n and "Base Rate" in n)

    return max(cand, key=score)


def find_template_cost(all_lanes: list[dict], cost_name: str) -> dict | None:
    for ln in all_lanes:
        for c in ln.get("Costs", []):
            if c.get("Cost") == cost_name:
                return copy.deepcopy(c)
    return None


def merged_tiered_fee_templates(
    lanes: list[dict], eqp: str, match_equipment
) -> list[dict]:
    """
    Union of matching tier blocks across all lanes, deduplicated by ``Cost`` name
    (first occurrence wins). For *DFT* merges, pass only ``_lanes_matching_rec_route`` so
    another route does not pollute tier donors.
    """
    seen: set[str] = set()
    out: list[dict] = []
    for ln in lanes:
        for c in ln.get("Costs", []):
            if not match_equipment(c.get("Cost") or "", eqp):
                continue
            name = (c.get("Cost") or "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            out.append(copy.deepcopy(c))
    return out


def _prototype_dft_fee_row(
    all_lanes: list[dict], match_equipment
) -> dict | None:
    """First *DFT* matching fee row on the card (for ``Rate by`` / ``Rule`` styling)."""
    cand = merged_tiered_fee_templates(all_lanes, "*DFT", match_equipment)
    return copy.deepcopy(cand[0]) if cand else None


def _strip_dft_baf_eu_ets_flat_columns(rec: dict) -> dict:
    """Copy without *DFT* BAF / EU ETS rate inputs (for batched new-route base costs)."""
    r = dict(rec)
    for col in DFT_BAF_EU_ETS_FLAT_COLS:
        r[col] = None
    return r


def _merged_dft_fee_templates_overlapping_window(
    all_lanes: list[dict],
    eff: date,
    exp: date,
    match_equipment,
) -> list[dict]:
    """
    Donor *DFT* fee tiers whose ``Validity period`` overlaps ``[eff, exp]`` —
    avoids pulling the entire historical chain when opening a new lane on one CSV window.
    """
    full = merged_tiered_fee_templates(all_lanes, "*DFT", match_equipment)
    return [
        copy.deepcopy(c)
        for c in full
        if _tier_validity_overlaps_window(c, eff, exp)
    ]


def _pick_dft_donor_tier_covering_union(
    donors: list[dict], union_lo: date, union_hi: date
) -> dict | None:
    """
    Pick the **tightest** donor tier whose validity fully contains ``[union_lo, union_hi]``.
    Used to duplicate one physical column (e.g. ``BAF Fee (since 15.01.2026)``) into
    several renamed columns with separate dates and prices.
    """
    candidates: list[tuple[dict, int]] = []
    for t in donors:
        d1, d2 = _tier_validity_bounds(t)
        if d1 is None:
            continue
        hi = d2 if d2 is not None else date(2099, 12, 31)
        if d1 <= union_lo and union_hi <= hi:
            span_days = (hi - d1).days
            candidates.append((t, span_days))
    if not candidates:
        return None

    def sort_key(item: tuple[dict, int]) -> tuple:
        t, span_days = item
        title = (t.get("Cost") or "").lower()
        # Among equal calendar spans, prefer the wide ``… (since …)`` column the user
        # splits in Excel (not a different band with the same end date).
        prefer_since = 0 if "since" in title else 1
        return (span_days, prefer_since)

    candidates.sort(key=sort_key)
    return copy.deepcopy(candidates[0][0])


def _collect_dft_rate_windows(
    sub_recs: list[dict], rate_col: str, ccy_col: str
) -> list[tuple[date, date, object, object]]:
    """Chronological (eff, exp, price, currency) per sub-row where ``rate_col`` is set."""
    rows: list[tuple[date, date, object, object]] = []
    for sr in sub_recs:
        if sr.get(rate_col) in (None, ""):
            continue
        eff = _parse_us_date(sr.get("RATE_EFFECTIVE_DATE__C"))
        exp = _parse_us_date(sr.get("RATE_EXPIRATION_DATE__C"))
        if not eff or not exp:
            continue
        rows.append((eff, exp, sr.get(rate_col), sr.get(ccy_col)))
    rows.sort(key=lambda x: (x[0], x[1]))
    return rows


def _dft_fee_emit_batch_split_clones(
    sub_recs: list[dict],
    all_lanes: list[dict],
    issues: list[str],
    *,
    fee_title: str,
    match_equipment,
    rate_col: str,
    ccy_col: str,
    issue_tag: str,
) -> list[dict]:
    """
    For a **new route** with several CSV fee windows: duplicate the single donor tier
    that spans the union of those windows (same column values as in Excel), rename each
    copy to the window range, set ``Validity period`` and CSV price — and **omit** that
    wide donor column from the output (no separate sibling historical tiers).
    """
    windows = _collect_dft_rate_windows(sub_recs, rate_col, ccy_col)
    if not windows:
        return []
    union_lo = min(w[0] for w in windows)
    union_hi = max(w[1] for w in windows)
    donors = merged_tiered_fee_templates(all_lanes, "*DFT", match_equipment)
    master = _pick_dft_donor_tier_covering_union(donors, union_lo, union_hi)
    if master is not None:
        out: list[dict] = []
        for eff, exp, pr, ccy in windows:
            row = copy.deepcopy(master)
            row["Cost"] = _tiered_fee_cost_label("*DFT", fee_title, eff, exp)
            row["Validity period"] = f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}"
            row["Currency"] = ccy
            row["Price"] = _price_num(pr)
            out.append(row)
        issues.append(
            f"  {issue_tag} *DFT: new-route batch — duplicated donor column "
            f"{master.get('Cost')!r} into {len(windows)} window column(s) "
            f"(values copied; wide tier not kept)."
        )
        return out
    # No single donor spans all windows — emit one minimal tier per window from prototype.
    proto = _prototype_dft_fee_row(all_lanes, match_equipment)
    rows_fb: list[dict] = []
    for eff, exp, pr, ccy in windows:
        rows_fb.extend(
            build_tiered_fee_block(
                [],
                "*DFT",
                eff,
                exp,
                ccy,
                pr,
                issues,
                fee_title=fee_title,
                match_equipment=match_equipment,
                issue_tag=issue_tag,
                dft_row_prototype=proto,
                dft_all_lanes=None,
            )
        )
    issues.append(
        f"  {issue_tag} *DFT: new-route batch — no single donor tier covers "
        f"{_fmt_dmy(union_lo)}–{_fmt_dmy(union_hi)}; using {len(windows)} standalone tier(s)."
    )
    return rows_fb


def build_lane_costs_for_new_route_batched(
    sub_recs: list[dict],
    donor_lane: dict,
    all_lanes: list[dict],
    issues: list[str],
) -> list[dict]:
    """
    One lane for several *DFT* fee windows: base/container costs from the first window
    row without *DFT* BAF/EU ETS rates; fee columns cloned from the donor ``since`` /
    wide tier per user split rules.
    """
    base_rec = _strip_dft_baf_eu_ets_flat_columns(sub_recs[0])
    out = build_lane_costs_from_update(base_rec, donor_lane, all_lanes, issues)
    out.extend(
        _dft_fee_emit_batch_split_clones(
            sub_recs,
            all_lanes,
            issues,
            fee_title="BAF Fee",
            match_equipment=baf_matches_equipment,
            rate_col="*DFTBAF_rate (per w/m)",
            ccy_col="*DFTBAF_Currency",
            issue_tag="BAF",
        )
    )
    out.extend(
        _dft_fee_emit_batch_split_clones(
            sub_recs,
            all_lanes,
            issues,
            fee_title="EU ETS Fee",
            match_equipment=eu_ets_matches_equipment,
            rate_col="*DFTEU_ETS_rate (per w/m)",
            ccy_col="*DFTEU_ETS_Currency",
            issue_tag="EU ETS",
        )
    )
    return out


def _missing_template_non_baf_cost(cost_label: str) -> dict:
    """
    No donor column pattern found for this ``Cost`` label — omit Rate by / Rule;
    Excel highlights the Rate-by cell in red (see ``missing_cost_template``).
    """
    return {
        "Cost": cost_label,
        "Validity period": "",
        "Cost to prolong": "",
        "Apply if": "Applies if invoiced by Carrier",
        "Rate by": "",
        "Rule": "",
        "Currency": None,
        "Price": None,
        "missing_cost_template": True,
    }


def _tiered_fee_cost_label(eqp: str, fee_title: str, eff: date, exp: date) -> str:
    """
    *DFT breakbulk tiers use date-only titles (same style as existing rate card columns).
    Container equipment keeps ``Fee (42P1 dd.mm.yyyy - dd.mm.yyyy)``.
    """
    if eqp == "*DFT":
        return f"{fee_title} ({_fmt_dmy(eff)} - {_fmt_dmy(exp)})"
    return f"{fee_title} ({eqp} {_fmt_dmy(eff)} - {_fmt_dmy(exp)})"


def _union_intervals_cover_range(
    intervals: list[tuple[date, date]], lo: date, hi: date
) -> bool:
    """True if merged contiguous intervals fully cover ``[lo, hi]``."""
    iv = [(a, b) for a, b in intervals if a <= b]
    if not iv:
        return False
    iv.sort(key=lambda x: x[0])
    cur_lo, cur_hi = iv[0]
    if cur_lo > lo:
        return False
    for a, b in iv[1:]:
        if a <= cur_hi + timedelta(days=1):
            cur_hi = max(cur_hi, b)
        else:
            return False
        if cur_hi >= hi:
            return True
    return cur_lo <= lo and cur_hi >= hi


def _lane_dft_union_covers_segment(
    lane: dict,
    seg_lo: date,
    seg_hi: date,
    fee_title: str,
    match_fn,
    target_price,
    target_ccy,
) -> bool:
    """
    *DFT tiers may split across several date-range columns; together they must cover
    ``[seg_lo, seg_hi]`` with matching price/currency on each overlapping tier.
    """
    clips: list[tuple[date, date]] = []
    for c in lane.get("Costs", []):
        name = c.get("Cost") or ""
        if not name.startswith(fee_title):
            continue
        if not match_fn(name, "*DFT"):
            continue
        d_lo, d_hi = _tier_validity_bounds(c)
        if d_lo is None:
            continue
        lo = max(seg_lo, d_lo)
        if d_hi is not None:
            hi = min(seg_hi, d_hi)
        else:
            hi = seg_hi
        if lo > hi:
            continue
        if _norm(c.get("Currency")) != _norm(target_ccy):
            return False
        if not _price_equal(c.get("Price"), target_price):
            return False
        clips.append((lo, hi))
    return _union_intervals_cover_range(clips, seg_lo, seg_hi)


def _dft_patch_overlapping_tier_rows(
    tiers: list[dict],
    eff: date,
    exp: date,
    currency,
    new_price,
) -> bool:
    """
    Update every tier row whose validity overlaps ``[eff, exp]`` if those clips
    union-cover the full window (no new Cost name).
    """
    clips_idx: list[tuple[date, date, int]] = []
    for i, t in enumerate(tiers):
        d_lo, d_hi = _tier_validity_bounds(t)
        if d_lo is None:
            continue
        lo = max(eff, d_lo)
        if d_hi is not None:
            hi = min(exp, d_hi)
        else:
            hi = exp
        if lo > hi:
            continue
        clips_idx.append((lo, hi, i))
    clips = [(a, b) for a, b, _ in clips_idx]
    if not _union_intervals_cover_range(clips, eff, exp):
        return False
    pr = _price_num(new_price)
    for _, _, i in clips_idx:
        tiers[i]["Currency"] = currency
        tiers[i]["Price"] = pr
    return True


def _baf_tier_covering_window(
    tiers: list[dict],
    eff: date,
    exp: date,
    eqp: str,
    *,
    fee_title: str | None = None,
    match_equipment=None,
) -> int | None:
    """
    Prefer a closed interval tier that contains ``[eff, exp]``; else an open-ended
    ``since`` tier that starts on/before ``eff``.

    For *DFT*, when several tiers contain the same window, prefer the **tightest**
    span; on an **equal** span, prefer a row whose **Validity period** indicates an open
    / ``since`` band (not inferred from the **Cost** name).

    When ``fee_title`` and ``match_equipment`` are passed (from :func:`build_tiered_fee_block`),
    bounds are taken from **Validity period** for all equipment codes.
    """
    best_key: tuple | None = None
    closed_i: int | None = None
    open_i: int | None = None
    open_d1: date | None = None
    for i, t in enumerate(tiers):
        if fee_title and match_equipment:
            d1, d2 = _fee_tier_semantic_open_since_bounds(
                t, eqp, fee_title, match_equipment
            )
        else:
            d1, d2 = _dft_semantic_validity_bounds(t, eqp)
        if d1 and d2:
            if d1 <= eff and exp <= d2:
                span = (d2 - d1).days
                if eqp == "*DFT":
                    vp = (t.get("Validity period") or "").lower()
                    prefer_open = "since" in vp or _tier_validity_bounds(t)[1] is None
                    if prefer_open:
                        key: tuple = (span, 0, i)
                    else:
                        key = (span, 1, i)
                else:
                    key = (span, i)
                if best_key is None or key < best_key:
                    best_key = key
                    closed_i = i
        elif d1 and not d2:
            if d1 <= eff:
                if open_d1 is None or d1 > open_d1:
                    open_d1 = d1
                    open_i = i
    return closed_i if closed_i is not None else open_i


def _dft_preferred_cost_title(
    all_lanes: list[dict],
    fee_title: str,
    match_fn,
    eqp: str,
    eff: date,
    exp: date,
) -> str | None:
    """
    If the rate card already has a *DFT column for the same start window (open-ended),
    matching **Validity period** start to ``eff``, reuse that **Cost** name.
    """
    if eqp != "*DFT":
        return None
    for ln in all_lanes:
        for c in ln.get("Costs", []):
            nm = c.get("Cost") or ""
            if not str(nm).startswith(fee_title) or not match_fn(nm, eqp):
                continue
            d1, d2 = _tier_validity_bounds(c)
            if d1 == eff and d2 == exp:
                return str(nm).strip()
            if _dft_breakbulk_since_style_validity(c) and d1 == eff and d2 is None:
                return str(nm).strip()
    return None


def _dft_cost_name_for_new_tier(
    all_lanes: list[dict] | None,
    eqp: str,
    fee_title: str,
    match_fn,
    eff: date,
    exp: date,
) -> str:
    if all_lanes is not None and eqp == "*DFT":
        alt = _dft_preferred_cost_title(all_lanes, fee_title, match_fn, eqp, eff, exp)
        if alt:
            return alt
    return _tiered_fee_cost_label(eqp, fee_title, eff, exp)


def _dft_apply_static_prototype(tier: dict, proto: dict | None) -> None:
    """Copy *DFT breakbulk field style from a sibling row (``Rate by``, ``Apply if``, etc.)."""
    if not proto:
        return
    for k in (
        "Apply if",
        "Rule",
        "Rate by",
    ):
        v = proto.get(k)
        if v is not None and str(v).strip():
            tier[k] = v


def dft_reuse_existing_subtier_rows_for_new_lane(
    all_lanes: list[dict],
    fee_title: str,
    match_fn,
    eff: date,
    exp: date,
    price,
    ccy,
    dft_row_prototype: dict | None,
) -> list[dict] | None:
    """
    For a **new** *DFT-only lane, if the card already has existing sub-tiers
    (e.g. ``EU ETS (01.01-14.01)`` and ``(15.01-31.03)``) fully inside
    ``[eff, exp]`` whose **union** covers the CSV window, return those
    **Cost** rows with the CSV **price** on each — instead of one combined
    ``(01.01-31.03)`` line. A wide row matching ``[eff, exp]`` exactly is
    skipped when at least two **strict** sub-tiers already partition the window.
    """
    if not all_lanes:
        return None
    by_name: dict[str, dict] = {}
    for ln in all_lanes:
        for c in ln.get("Costs", []):
            nm = c.get("Cost") or ""
            if not str(nm).startswith(fee_title) or not match_fn(nm, "*DFT"):
                continue
            d1, d2 = _tier_validity_bounds(c)
            if d1 is None or d2 is None:
                continue
            if d1 < eff or d2 > exp:
                continue
            k = str(nm).strip()
            if k not in by_name:
                by_name[k] = copy.deepcopy(c)
    if not by_name:
        return None
    entries: list[tuple[dict, date, date]] = []
    for t in by_name.values():
        d1, d2 = _tier_validity_bounds(t)
        if d1 and d2 and d1 >= eff and d2 <= exp:
            entries.append((t, d1, d2))
    entries.sort(key=lambda x: x[1])
    if not entries:
        return None

    def _ivs(el: list[tuple[dict, date, date]]) -> list[tuple[date, date]]:
        return [(a, b) for _t, a, b in el]

    no_full = [e for e in entries if not (e[1] == eff and e[2] == exp)]
    to_emit: list[tuple[dict, date, date]] | None = None
    if len(no_full) >= 2 and _union_intervals_cover_range(_ivs(no_full), eff, exp):
        to_emit = no_full
    elif len(no_full) == 0 and len(entries) >= 1:
        t0, d1, d2 = entries[0]
        if d1 == eff and d2 == exp:
            t2 = copy.deepcopy(t0)
            t2["Price"] = _price_num(price)
            t2["Currency"] = ccy
            if dft_row_prototype:
                _dft_apply_static_prototype(t2, dft_row_prototype)
            return [t2]
        return None
    if to_emit is None:
        return None
    out: list[dict] = []
    for t, _a, _b in to_emit:
        t2 = copy.deepcopy(t)
        t2["Price"] = _price_num(price)
        t2["Currency"] = ccy
        if dft_row_prototype:
            _dft_apply_static_prototype(t2, dft_row_prototype)
        out.append(t2)
    return out if out else None


def _dft_align_tier_cost_validity_to_csv_window(
    t: dict,
    tiers: list[dict],
    idx: int,
    fee_title: str,
    match_equipment,
    eff: date,
    exp: date,
    dft_all_lanes: list[dict] | None,
) -> None:
    """
    After :func:`_baf_tier_covering_window` in-place price update, set **Validity period** (and
    **Cost** when unambiguous) to this CSV ``[eff, exp]``.

    Without this, a ``since`` / wide tier keeps an older **Cost** title while receiving the next
    window's **Price** — :func:`merge_dft_fee_cost_rows_into_lane` then writes that price under
    the wrong column (e.g. 15.69 under ``BAF Fee (15.01.2026 - 14.04.2026)``).
    """
    preferred = _dft_preferred_cost_title(
        dft_all_lanes or [], fee_title, match_equipment, "*DFT", eff, exp
    )
    new_nm = preferred or _tiered_fee_cost_label("*DFT", fee_title, eff, exp)
    other_costs = {
        str(tiers[j].get("Cost") or "").strip()
        for j in range(len(tiers))
        if j != idx
    }
    if str(t.get("Cost") or "").strip() != new_nm.strip():
        if new_nm.strip() not in other_costs:
            t["Cost"] = new_nm
    t["Validity period"] = f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}"


def _dft_exact_validity_update_in_place(
    tiers: list[dict],
    eqp: str,
    eff: date,
    exp: date,
    fee_title: str,
    match_fn,
    currency,
    new_price,
) -> bool:
    """
    If a *DFT tier already uses the same calendar window as the CSV (including
    ``BAF Fee (since …)`` with full ``from … to …`` validity), update price on
    that row only — do not add a second column with a range-style title.
    """
    if eqp != "*DFT":
        return False
    for t in tiers:
        nm = t.get("Cost") or ""
        if not str(nm).startswith(fee_title) or not match_fn(nm, eqp):
            continue
        d1, d2 = _tier_validity_bounds(t)
        if d1 == eff and d2 == exp:
            t["Currency"] = currency
            t["Price"] = _price_num(new_price)
            return True
    return False


def build_tiered_fee_block(
    donor_costs: list[dict],
    eqp: str,
    eff: date,
    exp: date,
    currency,
    new_price,
    issues: list[str],
    *,
    fee_title: str,
    match_equipment,
    issue_tag: str,
    dft_row_prototype: dict | None = None,
    dft_all_lanes: list[dict] | None = None,
) -> list[dict]:
    tiers = [
        copy.deepcopy(c)
        for c in donor_costs
        if match_equipment(c.get("Cost") or "", eqp)
    ]
    if not tiers:
        if (
            eqp == "*DFT"
            and dft_all_lanes is not None
        ):
            reused = dft_reuse_existing_subtier_rows_for_new_lane(
                dft_all_lanes,
                fee_title,
                match_equipment,
                eff,
                exp,
                new_price,
                currency,
                dft_row_prototype,
            )
            if reused is not None:
                issues.append(
                    f"  {issue_tag} *DFT: reusing {len(reused)} existing sub-tier cost column(s) for "
                    f"window {_fmt_dmy(eff)}–{_fmt_dmy(exp)} (same price on each row)."
                )
                return reused
        issues.append(
            f"No {issue_tag} template in donor for equipment {eqp!r}; using single new tier."
        )
        if eqp == "*DFT":
            one = {
                "Cost": _dft_cost_name_for_new_tier(
                    dft_all_lanes, eqp, fee_title, match_equipment, eff, exp
                ),
                "Validity period": f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}",
                "Cost to prolong": "",
                "Apply if": "Applies if invoiced by Carrier",
                "Rate by": "Container/40FR",
                "Rule": "Regular rule",
                "Currency": currency,
                "Price": _price_num(new_price),
            }
            if dft_row_prototype:
                _dft_apply_static_prototype(one, dft_row_prototype)
        else:
            # New container BAF column: no prior rate-card row for this equipment — do not
            # invent Rate by / Rule (Excel marks Rate-by red via ``missing_cost_template``).
            one = {
                "Cost": _dft_cost_name_for_new_tier(
                    dft_all_lanes, eqp, fee_title, match_equipment, eff, exp
                ),
                "Validity period": f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}",
                "Cost to prolong": "",
                "Apply if": "Applies if invoiced by Carrier",
                "Rate by": "",
                "Rule": "",
                "Currency": currency,
                "Price": _price_num(new_price),
                "missing_cost_template": True,
            }
        return [one]

    cover_i = _baf_tier_covering_window(
        tiers, eff, exp, eqp, fee_title=fee_title, match_equipment=match_equipment
    )
    if cover_i is not None:
        t = tiers[cover_i]
        d_lo, d_hi = _fee_tier_semantic_open_since_bounds(
            t, eqp, fee_title, match_equipment
        )
        # Open-ended ``since`` tier + **short** finite CSV window → must split into a dated
        # Cost column + trailing ``since`` — do not overwrite the whole column here
        # (*DFT* + container ``since`` titles; Ocean Rates 20260327 / 20260306).
        carve_from_since = (
            d_lo is not None
            and d_hi is None
            and eff is not None
            and exp is not None
            and (exp - eff).days < DFT_SINCE_INPLACE_MIN_SPAN_DAYS
        )
        if not carve_from_since:
            t["Currency"] = currency
            t["Price"] = _price_num(new_price)
            if eqp == "*DFT":
                _dft_align_tier_cost_validity_to_csv_window(
                    t,
                    tiers,
                    cover_i,
                    fee_title,
                    match_equipment,
                    eff,
                    exp,
                    dft_all_lanes,
                )
            return tiers

    if eqp == "*DFT" and _dft_exact_validity_update_in_place(
        tiers, eqp, eff, exp, fee_title, match_equipment, currency, new_price
    ):
        return tiers

    # Short CSV window + ``since`` tier → must carve columns (split path), not patch overlaps.
    long_span = eff is not None and exp is not None and (
        (exp - eff).days >= DFT_SINCE_INPLACE_MIN_SPAN_DAYS
    )
    any_since_col = any(_dft_breakbulk_since_style_validity(t) for t in tiers)
    allow_overlap_patch = long_span or not any_since_col

    if (
        allow_overlap_patch
        and eqp == "*DFT"
        and _dft_patch_overlapping_tier_rows(tiers, eff, exp, currency, new_price)
    ):
        return tiers

    split_i = None
    for i, t in enumerate(tiers):
        d1, d2 = _fee_tier_semantic_open_since_bounds(t, eqp, fee_title, match_equipment)
        if d1 and d2 and d1 <= eff <= d2:
            split_i = i
            break
    if split_i is None:
        for i, t in enumerate(tiers):
            _, d2 = _fee_tier_semantic_open_since_bounds(t, eqp, fee_title, match_equipment)
            if d2 and eff <= d2:
                split_i = i
                break
    if split_i is None:
        for i, t in enumerate(tiers):
            d1, d2 = _fee_tier_semantic_open_since_bounds(t, eqp, fee_title, match_equipment)
            if d1 is not None and d2 is None and d1 <= eff:
                split_i = i
                break
    if split_i is None:
        split_i = len(tiers) - 1
        issues.append(
            f"{issue_tag} {eqp}: no tier contained {_fmt_dmy(eff)}; splitting last tier {tiers[split_i].get('Cost')!r}."
        )

    t_split = tiers[split_i]
    t_before = copy.deepcopy(t_split)
    d1, d2 = _fee_tier_semantic_open_since_bounds(
        t_split, eqp, fee_title, match_equipment
    )
    is_since = d1 is not None and d2 is None

    if is_since and eff == d1:
        # Long CSV window on ``since`` — single negotiated rate (see ``cover_i`` carve rule).
        if (
            eff is not None
            and exp is not None
            and (exp - eff).days >= DFT_SINCE_INPLACE_MIN_SPAN_DAYS
        ):
            t_up = copy.deepcopy(t_before)
            t_up["Currency"] = currency
            t_up["Price"] = _price_num(new_price)
            if dft_row_prototype:
                _dft_apply_static_prototype(t_up, dft_row_prototype)
            tiers[split_i] = t_up
            issues.append(
                f"  {issue_tag}: in-place price on {t_before.get('Cost')!r} "
                f"(CSV span {(exp - eff).days}d)."
            )
            return tiers
        # Replace one ``since`` column with (1) finite window matching the CSV and (2) a trailing
        # ``since`` starting the day after — duplicate currency/price from the former ``since``
        # row; per-lane prices follow merge rules for overlapping routes.
        t_win = copy.deepcopy(t_before)
        t_win["Cost"] = _tiered_fee_cost_label(eqp, fee_title, eff, exp)
        t_win["Validity period"] = f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}"
        t_win["Cost to prolong"] = t_before.get("Cost") or ""
        t_win["Currency"] = t_before.get("Currency")
        t_win["Price"] = _price_num(t_before.get("Price"))
        if dft_row_prototype:
            _dft_apply_static_prototype(t_win, dft_row_prototype)

        tail_start = exp + timedelta(days=1)
        t_tail = copy.deepcopy(t_before)
        t_tail["Cost"] = f"{fee_title} (since {_fmt_dmy(tail_start)})"
        t_tail["Validity period"] = f"from {_fmt_dmy(tail_start)}"
        t_tail["Cost to prolong"] = t_before.get("Cost") or ""
        t_tail["Currency"] = t_before.get("Currency")
        t_tail["Price"] = t_before.get("Price")
        if dft_row_prototype:
            _dft_apply_static_prototype(t_tail, dft_row_prototype)

        issues.append(
            f"  {issue_tag}: split ``since`` into window {_fmt_dmy(eff)}–{_fmt_dmy(exp)} "
            f"and trailing ``since {_fmt_dmy(tail_start)}`` (copied rate from prior column)."
        )
        return tiers[:split_i] + [t_win, t_tail] + tiers[split_i + 1 :]

    if is_since and eff > d1:
        left_hi = eff - timedelta(days=1)
        t_split["Validity period"] = f"from {_fmt_dmy(d1)} to {_fmt_dmy(left_hi)}"
        if eqp == "*DFT":
            t_split["Cost"] = f"{fee_title} ({_fmt_dmy(d1)}-{_fmt_dmy(left_hi)})"
        # Duplicate negotiated currency / unit rate from the carved ``since`` column into the left band.
        t_split["Currency"] = t_before.get("Currency")
        t_split["Price"] = _price_num(t_before.get("Price"))
        new_tier = copy.deepcopy(t_before)
        new_tier["Cost"] = _tiered_fee_cost_label(eqp, fee_title, eff, exp)
        new_tier["Validity period"] = f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}"
        new_tier["Cost to prolong"] = t_before.get("Cost") or ""
        new_tier["Currency"] = currency
        new_tier["Price"] = _price_num(new_price)
        if eqp == "*DFT" and dft_row_prototype:
            _dft_apply_static_prototype(new_tier, dft_row_prototype)
        return tiers[: split_i + 1] + [new_tier] + tiers[split_i + 1 :]

    if d1 and d2:
        t_split["Validity period"] = (
            f"from {_fmt_dmy(d1)} to {_fmt_dmy(eff - timedelta(days=1))}"
        )
        if eqp == "*DFT" and eff > d1:
            t_split["Cost"] = f"{fee_title} ({_fmt_dmy(d1)}-{_fmt_dmy(eff - timedelta(days=1))})"
    elif d2 and not d1:
        issues.append(
            f"{issue_tag} {eqp}: tier {t_split.get('Cost')!r} has open-ended validity; appending new tier only."
        )
    new_tier = copy.deepcopy(t_before)
    new_tier["Cost"] = _dft_cost_name_for_new_tier(
        dft_all_lanes, eqp, fee_title, match_equipment, eff, exp
    )
    new_tier["Validity period"] = f"from {_fmt_dmy(eff)} to {_fmt_dmy(exp)}"
    new_tier["Cost to prolong"] = t_before.get("Cost") or ""
    new_tier["Currency"] = currency
    new_tier["Price"] = _price_num(new_price)
    if eqp == "*DFT" and dft_row_prototype:
        _dft_apply_static_prototype(new_tier, dft_row_prototype)

    return tiers[: split_i + 1] + [new_tier] + tiers[split_i + 1 :]


def build_baf_block(
    donor_costs: list[dict],
    eqp: str,
    eff: date,
    exp: date,
    currency,
    new_price,
    issues: list[str],
) -> list[dict]:
    return build_tiered_fee_block(
        donor_costs,
        eqp,
        eff,
        exp,
        currency,
        new_price,
        issues,
        fee_title="BAF Fee",
        match_equipment=baf_matches_equipment,
        issue_tag="BAF",
    )


def build_lane_costs_from_update(
    rec: dict,
    donor_lane: dict,
    all_lanes: list[dict],
    issues: list[str],
) -> list[dict]:
    donor_costs = donor_lane.get("Costs", [])
    eff = _parse_us_date(rec.get("RATE_EFFECTIVE_DATE__C"))
    exp = _parse_us_date(rec.get("RATE_EXPIRATION_DATE__C"))
    if not eff or not exp:
        issues.append(f"Missing RATE_EFFECTIVE_DATE__C / RATE_EXPIRATION_DATE__C for {rec.get('KEY')!r}")

    eqs = equipment_in_update(rec)
    out: list[dict] = []
    kinds_order = ["BASE", "DTHC", "OTHC", "BAF"]

    for eq in eqs:
        eqp = eq_paren(eq)
        for kind in kinds_order:
            col_r = f"{eq}CNTR_{kind}_rate (per container)"
            if col_r not in rec or rec[col_r] in (None, ""):
                continue
            ccy = rec.get(f"{eq}CNTR_{kind}_Currency")
            pr = rec[col_r]
            if kind != "BAF":
                label = NON_BAF_KIND_LABEL[kind].format(eqp=eqp)
                tmpl = find_template_cost(all_lanes, label)
                if tmpl:
                    tmpl = copy.deepcopy(tmpl)
                else:
                    tmpl = _missing_template_non_baf_cost(label)
                    issues.append(
                        f"  no rate-card pattern for {label!r} — "
                        f"``Rate by``/``Rule`` left empty; mark Excel Rate-by cell red."
                    )
                tmpl["Currency"] = ccy
                tmpl["Price"] = _price_num(pr)
                out.append(tmpl)
            else:
                if eff and exp:
                    baf_src = merged_tiered_fee_templates(
                        all_lanes, eqp, baf_matches_equipment
                    )
                    if not baf_src:
                        baf_src = [
                            c
                            for c in donor_costs
                            if baf_matches_equipment(c.get("Cost") or "", eqp)
                        ]
                        baf_src = copy.deepcopy(baf_src)
                    full_baf = build_baf_block(
                        baf_src, eqp, eff, exp, ccy, pr, issues
                    )
                    filt_baf = _filter_container_baf_block_to_csv_window(
                        full_baf, eqp, eff, exp
                    )
                    if len(filt_baf) < len(full_baf):
                        issues.append(
                            f"  BAF {eqp}: removed {len(full_baf) - len(filt_baf)} donor tier "
                            f"row(s) outside CSV window {_fmt_dmy(eff)}–{_fmt_dmy(exp)} "
                            f"(historical prices not in this update)."
                        )
                    out.extend(filt_baf if filt_baf else full_baf)
                    if not filt_baf and full_baf:
                        issues.append(
                            f"  BAF {eqp}: CSV-window filter yielded no rows; kept full donor chain."
                        )

    route_for_dft = _lanes_matching_rec_route(all_lanes, rec)
    lane_pool = route_for_dft if route_for_dft else all_lanes
    new_route_no_lane = not bool(route_for_dft)
    if eff and exp:
        if rec.get("*DFTBAF_rate (per w/m)") not in (None, ""):
            if new_route_no_lane:
                baf_src = _merged_dft_fee_templates_overlapping_window(
                    all_lanes, eff, exp, baf_matches_equipment
                )
                proto_b = _prototype_dft_fee_row(all_lanes, baf_matches_equipment)
            else:
                baf_src = merged_tiered_fee_templates(
                    lane_pool, "*DFT", baf_matches_equipment
                )
                if not baf_src:
                    baf_src = copy.deepcopy(
                        [
                            c
                            for c in donor_costs
                            if baf_matches_equipment(c.get("Cost") or "", "*DFT")
                        ]
                    )
                proto_b = baf_src[0] if baf_src else None
            out.extend(
                build_tiered_fee_block(
                    baf_src,
                    "*DFT",
                    eff,
                    exp,
                    rec.get("*DFTBAF_Currency"),
                    rec.get("*DFTBAF_rate (per w/m)"),
                    issues,
                    fee_title="BAF Fee",
                    match_equipment=baf_matches_equipment,
                    issue_tag="BAF",
                    dft_row_prototype=proto_b,
                    dft_all_lanes=lane_pool,
                )
            )
        if rec.get("*DFTEU_ETS_rate (per w/m)") not in (None, ""):
            if new_route_no_lane:
                ets_src = _merged_dft_fee_templates_overlapping_window(
                    all_lanes, eff, exp, eu_ets_matches_equipment
                )
                proto_e = _prototype_dft_fee_row(all_lanes, eu_ets_matches_equipment)
            else:
                ets_src = merged_tiered_fee_templates(
                    lane_pool, "*DFT", eu_ets_matches_equipment
                )
                if not ets_src:
                    ets_src = copy.deepcopy(
                        [
                            c
                            for c in donor_costs
                            if eu_ets_matches_equipment(c.get("Cost") or "", "*DFT")
                        ]
                    )
                proto_e = ets_src[0] if ets_src else None
            out.extend(
                build_tiered_fee_block(
                    ets_src,
                    "*DFT",
                    eff,
                    exp,
                    rec.get("*DFTEU_ETS_Currency"),
                    rec.get("*DFTEU_ETS_rate (per w/m)"),
                    issues,
                    fee_title="EU ETS Fee",
                    match_equipment=eu_ets_matches_equipment,
                    issue_tag="EU ETS",
                    dft_row_prototype=proto_e,
                    dft_all_lanes=lane_pool,
                )
            )

    out.extend(
        build_dft_breakbulk_wm_base_cost_rows(
            rec, all_lanes, issues, validity_eff_exp=None
        )
    )
    return out


def rec_has_cntr_rate_column_values(rec: dict) -> bool:
    """True if any ``42P1CNTR_*`` style container rate column has a value."""
    for col in rec:
        rk = rate_column_eq_kind(col)
        if rk and rec.get(col) not in (None, ""):
            return True
    return False


def rec_has_dft_rate_column_values(rec: dict) -> bool:
    """True if any ``*DFTBAF_*`` / ``*DFTEU_ETS_*`` rate column has a value (flat or in windows)."""
    wins = rec.get(DFT_BAF_EU_ETS_WINDOWS_KEY)
    if isinstance(wins, list):
        for w in wins:
            if not isinstance(w, dict):
                continue
            for col in DFT_BAF_EU_ETS_FLAT_COLS:
                if col.endswith("_rate (per w/m)") and w.get(col) not in (None, ""):
                    return True
    for col in rec:
        if dft_rate_column_kind(col) and rec.get(col) not in (None, ""):
            return True
    return False


def rec_has_dft_baf_fee_values(rec: dict) -> bool:
    """True if ``rec`` carries a non-empty *DFT* BAF w/m rate (table or ``DFT_BAF_EU_ETS_WINDOWS``)."""
    wins = rec.get(DFT_BAF_EU_ETS_WINDOWS_KEY)
    if isinstance(wins, list):
        for w in wins:
            if not isinstance(w, dict):
                continue
            if w.get("*DFTBAF_rate (per w/m)") not in (None, ""):
                return True
    for col in rec:
        k = dft_rate_column_kind(col)
        if k and k[1] == "BAF" and rec.get(col) not in (None, ""):
            return True
    return False


def rec_has_dft_eu_ets_fee_values(rec: dict) -> bool:
    """True if ``rec`` carries a non-empty *DFT* EU ETS w/m rate (table or windows)."""
    wins = rec.get(DFT_BAF_EU_ETS_WINDOWS_KEY)
    if isinstance(wins, list):
        for w in wins:
            if not isinstance(w, dict):
                continue
            if w.get("*DFTEU_ETS_rate (per w/m)") not in (None, ""):
                return True
    for col in rec:
        k = dft_rate_column_kind(col)
        if k and k[1] == "EU_ETS" and rec.get(col) not in (None, ""):
            return True
    return False


def merged_records_touch_dft_baf_fee(merged_records: list[dict]) -> bool:
    """True if any combined CSV row for this merge includes *DFT* BAF fee data."""
    for rec in merged_records:
        for sr in flat_records_from_combined_merge_rec(rec):
            if rec_has_dft_baf_fee_values(sr):
                return True
    return False


def merged_records_touch_dft_eu_ets_fee(merged_records: list[dict]) -> bool:
    """True if any combined CSV row for this merge includes *DFT* EU ETS fee data."""
    for rec in merged_records:
        for sr in flat_records_from_combined_merge_rec(rec):
            if rec_has_dft_eu_ets_fee_values(sr):
                return True
    return False


def _lane_validity_overlaps_csv_window(
    ln: dict, upd_lo: date, upd_hi: date
) -> bool:
    """Lane ``Valid from`` / ``Valid to`` intersects ``[upd_lo, upd_hi]`` (day granularity)."""
    w = _lane_validity_parsed(ln)
    if not w:
        return False
    lo, hi = w
    return _validity_non_disjoint(lo, hi, upd_lo, upd_hi)


def next_lane_numbers(lanes: list[dict], count: int) -> list[int]:
    mx = 0
    for ln in lanes:
        try:
            mx = max(mx, int(str(ln.get("Lane #", "")).strip()))
        except ValueError:
            continue
    return list(range(mx + 1, mx + 1 + count))


def match_lane(
    lanes: list[dict], carrier: str, svc: str, origin: str, dest: str
) -> dict | None:
    for ln in lanes:
        if (
            _norm(ln.get("Carrier")) == _norm(carrier)
            and _norm(ln.get("Service")) == _norm(svc)
            and _csv_port_matches_lane_field(origin, ln.get("Origin Port") or "")
            and _csv_port_matches_lane_field(dest, ln.get("Destination Port") or "")
        ):
            return ln
    return None


def lanes_matching_route(
    lanes: list[dict], carrier: str, svc: str, origin: str, dest: str
) -> list[dict]:
    """All lanes with same Carrier / Service / origin / destination (port aliases allowed)."""
    out: list[dict] = []
    for ln in lanes:
        if (
            _norm(ln.get("Carrier")) == _norm(carrier)
            and _norm(ln.get("Service")) == _norm(svc)
            and _csv_port_matches_lane_field(origin, ln.get("Origin Port") or "")
            and _csv_port_matches_lane_field(dest, ln.get("Destination Port") or "")
        ):
            out.append(ln)
    return out


def _lanes_matching_rec_route(lanes: list[dict], rec: dict) -> list[dict]:
    """
    Lanes on the same trade route as the update row. Used to build *DFT* BAF/EU ETS tier
    chains so another route’s columns (e.g. an older lane still on ``since 15.01``) are not
    pulled into ``merged_tiered_fee_templates``.
    """
    car = _norm(rec.get("CARRIER"))
    svc = service_cy_cy(rec.get("SERVICE__C", ""))
    origin = _norm(rec.get("ORIGIN_LOCATION_NAME__C"))
    dest = _norm(rec.get("DESTINATION_LOCATION_NAME__C"))
    if not car or not svc or not origin or not dest:
        return []
    return lanes_matching_route(lanes, car, svc, origin, dest)


def _date_range_intersection(
    a0: date, a1: date, b0: date, b1: date
) -> tuple[date, date] | None:
    s = max(a0, b0)
    e = min(a1, b1)
    if s > e:
        return None
    return (s, e)


def _tier_covers_segment(
    d_lo: date | None, d_hi: date | None, seg_lo: date, seg_hi: date
) -> bool:
    """Tier validity (closed or open-ended) fully contains ``[seg_lo, seg_hi]``."""
    if d_lo is None:
        return False
    if d_hi is not None:
        return d_lo <= seg_lo and seg_hi <= d_hi
    return d_lo <= seg_lo


def _price_equal(a, b) -> bool:
    pa, pb = _price_num(a), _price_num(b)
    if pa is None and pb is None:
        return True
    if pa is None or pb is None:
        return False
    try:
        return abs(float(pa) - float(pb)) < 1e-9
    except (TypeError, ValueError):
        return str(pa) == str(pb)


def _lane_has_covering_tiered_fee(
    lane: dict,
    seg_lo: date,
    seg_hi: date,
    eqp: str,
    fee_title: str,
    match_fn,
    target_price,
    target_ccy,
) -> bool:
    for c in lane.get("Costs", []):
        name = c.get("Cost") or ""
        if not name.startswith(fee_title):
            continue
        if not match_fn(name, eqp):
            continue
        d_lo, d_hi = _tier_validity_bounds(c)
        if not _tier_covers_segment(d_lo, d_hi, seg_lo, seg_hi):
            continue
        if _norm(c.get("Currency")) != _norm(target_ccy):
            continue
        if not _price_equal(c.get("Price"), target_price):
            continue
        return True
    return False


def rec_has_non_tiered_rate_values(rec: dict) -> bool:
    """BASE / DTHC / OTHC etc. — skip logic does not handle these yet."""
    for col in rec:
        rk = rate_column_eq_kind(col)
        if rk and rk[1] != "BAF" and rec.get(col) not in (None, ""):
            return True
    return False


def rec_non_tiered_rate_columns(rec: dict) -> list[str]:
    out: list[str] = []
    for col in rec:
        rk = rate_column_eq_kind(col)
        if rk and rk[1] != "BAF" and rec.get(col) not in (None, ""):
            out.append(col)
    return out


def _explain_dft_union_miss(
    lane: dict,
    seg_lo: date,
    seg_hi: date,
    fee_title: str,
    match_fn,
    target_price,
    target_ccy,
) -> str:
    lids = _norm(lane.get("Lane #"))
    clips: list[tuple[date, date]] = []
    problems: list[str] = []
    for c in lane.get("Costs", []):
        name = c.get("Cost") or ""
        if not name.startswith(fee_title):
            continue
        if not match_fn(name, "*DFT"):
            continue
        d_lo, d_hi = _tier_validity_bounds(c)
        if d_lo is None:
            continue
        lo = max(seg_lo, d_lo)
        hi = min(seg_hi, d_hi if d_hi else seg_hi)
        if d_hi is None:
            hi = seg_hi
        if lo > hi:
            continue
        if _norm(c.get("Currency")) != _norm(target_ccy):
            problems.append(f"{name!r} ccy={c.get('Currency')!r} want {target_ccy!r}")
            continue
        if not _price_equal(c.get("Price"), target_price):
            problems.append(f"{name!r} price={c.get('Price')!r} want {target_price!r}")
            continue
        clips.append((lo, hi))
    if problems:
        return f"lane {lids}: " + "; ".join(problems[:5])
    if not _union_intervals_cover_range(clips, seg_lo, seg_hi):
        return (
            f"lane {lids}: *DFT {fee_title} tiers (date-title columns) do not jointly cover "
            f"{_fmt_dmy(seg_lo)}–{_fmt_dmy(seg_hi)}"
        )
    return f"lane {lids}: union check failed"


def _explain_lane_tier_miss(
    lane: dict,
    seg_lo: date,
    seg_hi: date,
    eqp: str,
    fee_title: str,
    match_fn,
    target_price,
    target_ccy,
) -> str:
    """One short reason why no covering tier was found."""
    lids = _norm(lane.get("Lane #"))
    tried: list[str] = []
    for c in lane.get("Costs", []):
        name = c.get("Cost") or ""
        if not name.startswith(fee_title):
            continue
        if not match_fn(name, eqp):
            tried.append(f"{name!r}: wrong equipment matcher for eqp={eqp!r}")
            continue
        d_lo, d_hi = _tier_validity_bounds(c)
        if not _tier_covers_segment(d_lo, d_hi, seg_lo, seg_hi):
            tried.append(
                f"{name!r}: validity {_fmt_dmy(d_lo) if d_lo else '?'}–"
                f"{_fmt_dmy(d_hi) if d_hi else 'open'} does not cover "
                f"{_fmt_dmy(seg_lo)}–{_fmt_dmy(seg_hi)}"
            )
            continue
        if _norm(c.get("Currency")) != _norm(target_ccy):
            tried.append(
                f"{name!r}: currency {c.get('Currency')!r} vs target {target_ccy!r}"
            )
            continue
        if not _price_equal(c.get("Price"), target_price):
            tried.append(
                f"{name!r}: price {c.get('Price')!r} vs target {target_price!r}"
            )
            continue
    if not tried:
        return f"lane {lids}: no {fee_title} row for eqp={eqp!r}"
    return f"lane {lids}: " + "; ".join(tried[:4])


def merge_record_redundant_with_existing_lanes_explain(
    rec: dict, lanes: list[dict]
) -> tuple[bool, list[str]]:
    """
    Returns whether this CSV row is redundant (tiered fees already satisfied on route lanes)
    plus trace lines for the merge report.
    """
    t: list[str] = []
    t.append("  redundant-check: start")

    if not lanes:
        t.append("  redundant-check: FAIL — rate card has zero lanes")
        return False, t

    ntd = rec_non_tiered_rate_columns(rec)
    if ntd:
        t.append(
            "  redundant-check: FAIL — non-tiered rate columns (BASE/DTHC/OTHC) have "
            f"values: {ntd!r} (auto-skip only handles tiered BAF / EU ETS)"
        )
        return False, t

    upd_from = _parse_us_date(rec.get("RATE_EFFECTIVE_DATE__C"))
    upd_to = _parse_us_date(rec.get("RATE_EXPIRATION_DATE__C"))
    if not upd_from or not upd_to:
        t.append(
            "  redundant-check: FAIL — missing RATE_EFFECTIVE_DATE__C or RATE_EXPIRATION_DATE__C"
        )
        return False, t

    car = _norm(rec.get("CARRIER"))
    svc = service_cy_cy(rec.get("SERVICE__C", ""))
    origin = _norm(rec.get("ORIGIN_LOCATION_NAME__C"))
    dest = _norm(rec.get("DESTINATION_LOCATION_NAME__C"))
    t.append(
        f"  redundant-check: route match keys — Carrier={car!r} Service={svc!r} "
        f"Origin={origin!r} Dest={dest!r}"
    )
    t.append(
        f"  redundant-check: CSV window {_fmt_dmy(upd_from)}–{_fmt_dmy(upd_to)} "
        f"(from RATE_EFFECTIVE_DATE__C / RATE_EXPIRATION_DATE__C)"
    )

    route_lanes = lanes_matching_route(lanes, car, svc, origin, dest)
    if not route_lanes:
        t.append(
            "  redundant-check: FAIL — no lane with same Carrier+Service+Origin+Dest "
            "(port aliases like CNSHA/CNSHG allowed; compare JSON lane Service / ports "
            "to SERVICE__C and *_LOCATION_NAME__C)"
        )
        sample = []
        for ln in lanes[:5]:
            sample.append(
                f"Lane#{ln.get('Lane #')} svc={_norm(ln.get('Service'))!r} "
                f"o={_norm(ln.get('Origin Port'))!r} d={_norm(ln.get('Destination Port'))!r}"
            )
        if sample:
            t.append("  redundant-check: first lanes for contrast: " + " | ".join(sample))
        return False, t

    t.append(
        "  redundant-check: matching lanes — "
        + ", ".join(f"{_norm(x.get('Lane #'))}" for x in route_lanes)
    )

    checks: list[tuple[str, str, object, object, object]] = []

    if rec.get("*DFTBAF_rate (per w/m)") not in (None, ""):
        checks.append(
            (
                "*DFT",
                "BAF Fee",
                baf_matches_equipment,
                rec.get("*DFTBAF_rate (per w/m)"),
                rec.get("*DFTBAF_Currency"),
            )
        )
    if rec.get("*DFTEU_ETS_rate (per w/m)") not in (None, ""):
        checks.append(
            (
                "*DFT",
                "EU ETS Fee",
                eu_ets_matches_equipment,
                rec.get("*DFTEU_ETS_rate (per w/m)"),
                rec.get("*DFTEU_ETS_Currency"),
            )
        )

    for col in rec:
        rk = rate_column_eq_kind(col)
        if not rk or rk[1] != "BAF":
            continue
        if rec.get(col) in (None, ""):
            continue
        eq, _ = rk
        eqp = eq_paren(eq)
        checks.append(
            (
                eqp,
                "BAF Fee",
                baf_matches_equipment,
                rec[col],
                rec.get(f"{eq}CNTR_BAF_Currency"),
            )
        )

    if not checks:
        t.append(
            "  redundant-check: FAIL — no tiered columns to evaluate "
            "(*DFTBAF_rate, *DFTEU_ETS_rate, or *CNTR_BAF_rate)"
        )
        return False, t

    for eqp, fee_title, match_fn, price, ccy in checks:
        t.append(
            f"  redundant-check: tier — eqp={eqp!r} {fee_title} price={price!r} ccy={ccy!r}"
        )
        any_overlap = False
        for ln in route_lanes:
            lf = _parse_dmy_text(_norm(ln.get("Valid from")))
            lt = _parse_dmy_text(_norm(ln.get("Valid to")))
            lids = _norm(ln.get("Lane #"))
            if not lf or not lt:
                t.append(
                    f"  redundant-check: lane {lids} skipped — invalid Valid from/to "
                    f"({ln.get('Valid from')!r} / {ln.get('Valid to')!r})"
                )
                continue
            seg = _date_range_intersection(upd_from, upd_to, lf, lt)
            if seg is None:
                t.append(
                    f"  redundant-check: lane {lids} no date overlap "
                    f"(lane {_fmt_dmy(lf)}–{_fmt_dmy(lt)} vs update)"
                )
                continue
            any_overlap = True
            t.append(
                f"  redundant-check: lane {lids} segment {_fmt_dmy(seg[0])}–{_fmt_dmy(seg[1])} "
                f"(lane window ∩ CSV window)"
            )
            if eqp == "*DFT":
                satisfied = _lane_dft_union_covers_segment(
                    ln, seg[0], seg[1], fee_title, match_fn, price, ccy
                )
            else:
                satisfied = _lane_has_covering_tiered_fee(
                    ln,
                    seg[0],
                    seg[1],
                    eqp,
                    fee_title,
                    match_fn,
                    price,
                    ccy,
                )
            if satisfied:
                t.append(
                    f"  redundant-check: lane {lids} OK — "
                    f"{'union of *DFT tiers' if eqp == '*DFT' else 'covering tier'} "
                    f"+ price/ccy match"
                )
            else:
                why = (
                    _explain_dft_union_miss(
                        ln, seg[0], seg[1], fee_title, match_fn, price, ccy
                    )
                    if eqp == "*DFT"
                    else _explain_lane_tier_miss(
                        ln, seg[0], seg[1], eqp, fee_title, match_fn, price, ccy
                    )
                )
                t.append(f"  redundant-check: FAIL — {why}")
                return False, t
        if not any_overlap:
            t.append(
                "  redundant-check: FAIL — CSV window does not overlap any matching lane validity"
            )
            return False, t

    t.append("  redundant-check: PASS — would not change tiered fees; skipping merge for this row")
    return True, t


def strip_equipment_costs(lane_costs: list[dict], eqp: str) -> list[dict]:
    return [c for c in lane_costs if not cost_belongs_equipment(c.get("Cost") or "", eqp)]


def snapshot_costs_with_replaced_equipment(
    snapshot_costs: list[dict],
    rec: dict,
    new_costs: list[dict],
) -> list[dict]:
    """Full copy of lane costs, with CSV equipment blocks replaced by ``new_costs``."""
    mc = copy.deepcopy(snapshot_costs)
    for eqp in sorted(
        {eq_paren(e) for e in equipment_in_update(rec)},
        key=eq_sort_key,
    ):
        mc = strip_equipment_costs(mc, eqp)
    mc.extend(copy.deepcopy(new_costs))
    return mc


def filter_dft_fee_cost_rows_for_csv_window(
    new_costs: list[dict], sr: dict
) -> list[dict]:
    """
    Subset of ``new_costs`` that belong to **this** CSV window only: *DFT* BAF or EU ETS rows
    whose ``Validity period`` matches the window dates and whose ``Price`` matches the CSV rate
    for that fee. Excludes full ``build_tiered_fee_block`` donor chains (historical BAF columns
    like 21.31) so we do not overwrite unrelated tier cells on existing lanes.
    """
    eff = _parse_us_date(sr.get("RATE_EFFECTIVE_DATE__C"))
    exp = _parse_us_date(sr.get("RATE_EXPIRATION_DATE__C"))
    if not eff or not exp:
        return []
    pr_baf = sr.get("*DFTBAF_rate (per w/m)")
    pr_ets = sr.get("*DFTEU_ETS_rate (per w/m)")
    out: list[dict] = []
    for c in new_costs:
        nm = str(c.get("Cost") or "")
        if baf_matches_equipment(nm, "*DFT"):
            if pr_baf in (None, "") or not _numeric_fee_equal(c.get("Price"), pr_baf):
                continue
        elif eu_ets_matches_equipment(nm, "*DFT"):
            if pr_ets in (None, "") or not _numeric_fee_equal(c.get("Price"), pr_ets):
                continue
        else:
            continue
        if not _tier_validity_overlaps_window(c, eff, exp):
            continue
        out.append(copy.deepcopy(c))
    return out


def _dft_breakbulk_tier_intersects_lane_validity(c: dict, ln: dict) -> bool:
    """
    True if the cost row's ``Validity period`` overlaps the lane's ``Valid from`` / ``Valid to``.
    """
    lw = _lane_validity_parsed(ln)
    if not lw:
        return True
    lane_lo, lane_hi = lw
    d1, d2 = _tier_validity_bounds(c)
    if d1 is not None and d2 is not None:
        return _date_range_intersection(lane_lo, lane_hi, d1, d2) is not None
    if d1 is not None and d2 is None:
        return d1 <= lane_hi
    if d1 is None and d2 is not None:
        return lane_lo <= d2
    return True


def filter_dft_fee_cost_rows_for_lane(patch: list[dict], ln: dict) -> list[dict]:
    """
    Drop tier rows whose validity does **not** intersect the lane's ``Valid from`` / ``Valid to``.
    A CSV fee band may sit outside a lane's service period (e.g. ``15.04.2026–31.10.2026`` vs lane
    ending ``06.03.2026``) — those rows must not be merged onto that lane.
    """
    if not _lane_validity_parsed(ln):
        return list(patch)
    return [c for c in patch if _dft_breakbulk_tier_intersects_lane_validity(c, ln)]


def prune_dft_breakbulk_fees_outside_lane_window(lane: dict) -> None:
    """
    Remove *DFT* breakbulk BAF / EU ETS cost rows whose calendar band does not intersect the lane
    service window.

    Card-wide :func:`_migrate_since_tier_to_split_columns` deep-copies the ``since`` row into **each**
    finite split column for **every** lane, so a lane that ends before e.g. ``15.04.2026`` still gets
    ``BAF Fee (15.04.2026 - 31.10.2026)`` with a **duplicate** price from the former ``since`` row.
    """
    costs = lane.get("Costs")
    if not costs:
        return
    kept: list[dict] = []
    for c in costs:
        nm = str(c.get("Cost") or "")
        if not (
            baf_matches_equipment(nm, "*DFT")
            or eu_ets_matches_equipment(nm, "*DFT")
        ):
            kept.append(c)
            continue
        if _dft_breakbulk_tier_intersects_lane_validity(c, lane):
            kept.append(c)
    lane["Costs"] = kept


def merge_dft_fee_cost_rows_into_lane(lane: dict, new_costs: list[dict]) -> None:
    """
    Apply CSV-built tier rows to one lane: same ``Cost`` name → replace row; new name → append.
    Used when an update row is *DFT* BAF/EU ETS only so every route lane that overlaps the CSV
    window gets that window’s prices (not only the first overlapping lane).
    """
    costs = lane.get("Costs")
    if costs is None:
        costs = []
        lane["Costs"] = costs
    by_name: dict[str, int] = {}
    for i, c in enumerate(costs):
        nm = str(c.get("Cost") or "").strip()
        if nm and nm not in by_name:
            by_name[nm] = i
    for nc in new_costs:
        nm = str(nc.get("Cost") or "").strip()
        if not nm:
            continue
        row = copy.deepcopy(nc)
        if nm in by_name:
            costs[by_name[nm]] = row
        else:
            costs.append(row)
            by_name[nm] = len(costs) - 1


def _dedupe_duplicate_cost_names_on_lane(costs: list[dict]) -> list[dict]:
    """Keep the first cost row for each non-empty ``Cost`` name."""
    if not costs:
        return costs
    seen: set[str] = set()
    out: list[dict] = []
    for c in costs:
        nm = str(c.get("Cost") or "").strip()
        if nm:
            if nm in seen:
                continue
            seen.add(nm)
        out.append(c)
    return out


def _lane_route_key(ln: dict) -> tuple:
    return (
        _norm(ln.get("Carrier")),
        _norm(ln.get("Service")),
        _norm(ln.get("Origin Port") or ""),
        _norm(ln.get("Destination Port") or ""),
    )


def _int_lane_num(ln: dict) -> int:
    try:
        return int(str(ln.get("Lane #", "")).strip() or 0)
    except ValueError:
        return 0


def _lane_validity_parsed(
    ln: dict,
) -> tuple[date, date] | None:
    a = _parse_dmy_text(_norm(ln.get("Valid from")))
    b = _parse_dmy_text(_norm(ln.get("Valid to")))
    if not a or not b:
        return None
    return a, b


def _validity_non_disjoint(
    a1: date, a2: date, b1: date, b2: date
) -> bool:
    return a1 <= b2 and b1 <= a2


def _uf_find(parent: dict, x: int) -> int:
    while parent[x] != x:
        parent[x] = parent[parent[x]]
        x = parent[x]
    return x


def _uf_union(parent: dict, x: int, y: int) -> None:
    rx, ry = _uf_find(parent, x), _uf_find(parent, y)
    if rx != ry:
        parent[rx] = ry


def _merge_dft_tiered_lane_group(lns: list[dict]) -> dict:
    lns = sorted(lns, key=_int_lane_num)
    out = copy.deepcopy(lns[0])
    d_lo: date | None = None
    d_hi: date | None = None
    for x in lns:
        w = _lane_validity_parsed(x)
        if w:
            lo, hi = w
            d_lo = lo if d_lo is None else min(d_lo, lo)
            d_hi = hi if d_hi is None else max(d_hi, hi)
    if d_lo and d_hi:
        out["Valid from"] = _fmt_dmy(d_lo)
        out["Valid to"] = _fmt_dmy(d_hi)
    seen: dict[str, dict] = {}
    for x in lns:
        for c in x.get("Costs", []):
            name = c.get("Cost")
            if not name:
                continue
            k = str(name).strip()
            if k not in seen:
                seen[k] = copy.deepcopy(c)
    out["Costs"] = [seen[k] for k in sort_rate_card_cost_names(list(seen.keys()))]
    return out


def consolidate_mergeable_route_lanes(
    lanes: list[dict], issues: list[str] | None
) -> list[dict]:
    """
    Merge lanes that share the same route, have **overlapping** lane validity
    windows, and **no conflicting** ``Cost`` lines (same name must have same
    currency/price). Lane dates become min–max; costs are unioned by name.

    Previously only *DFT-only* breakbulk lanes were merged; any compatible lane
    pair (e.g. appended lanes that differ only by tier columns) can merge.
    """
    log.debug(
        "consolidate_mergeable_route_lanes: %s lane(s) in — grouping by route + overlapping validity.",
        len(lanes),
    )
    by_route: defaultdict[tuple, list[int]] = defaultdict(list)
    for i, ln in enumerate(lanes):
        if _lane_validity_parsed(ln) is None:
            continue
        by_route[_lane_route_key(ln)].append(i)

    merged_leaders: dict[int, dict] = {}
    merged_slaves: set[int] = set()
    for _rkey, mem in by_route.items():
        if len(mem) < 2:
            continue
        parent: dict[int, int] = {i: i for i in mem}
        for a in mem:
            for b in mem:
                if a >= b:
                    continue
                w1 = _lane_validity_parsed(lanes[a])
                w2 = _lane_validity_parsed(lanes[b])
                if not w1 or not w2:
                    continue
                a1, a2 = w1
                b1, b2 = w2
                if not _validity_non_disjoint(a1, a2, b1, b2):
                    continue
                cm_a = _lane_cost_map(lanes[a])
                cm_b = _lane_cost_map(lanes[b])
                if not set(cm_a) & set(cm_b):
                    # Overlapping validity + disjoint cost columns (e.g. 45G0 vs 42UTH) must
                    # stay separate lanes; do not union into one row.
                    continue
                if not _cost_maps_mergeable(cm_a, cm_b):
                    continue
                _uf_union(parent, a, b)
        by_root: defaultdict[int, list[int]] = defaultdict(list)
        for m in mem:
            by_root[_uf_find(parent, m)].append(m)
        for comp in by_root.values():
            if len(comp) < 2:
                continue
            comp_lns = [lanes[i] for i in comp]
            merged = _merge_dft_tiered_lane_group(comp_lns)
            leader = min(comp, key=lambda j: _int_lane_num(lanes[j]))
            merged["Lane #"] = lanes[leader].get("Lane #")
            merged_leaders[leader] = merged
            for j in comp:
                if j != leader:
                    merged_slaves.add(j)

    if not merged_leaders:
        return lanes
    if issues is not None:
        issues.append(
            f"  post-merge: consolidated {len(merged_slaves)} compatible overlapping "
            f"route lane(s) into {len(merged_leaders)} lane(s) (lane window = min–max; "
            f"cost names deduped; same name must have matched currency/price)."
        )
    out: list[dict] = []
    for i, ln in enumerate(lanes):
        if i in merged_slaves:
            continue
        if i in merged_leaders:
            out.append(merged_leaders[i])
        else:
            out.append(ln)
    return out


def _migrate_since_tier_to_split_columns(
    lanes: list[dict],
    issues: list[str],
    *,
    since_cost: str,
    split_specs: tuple[tuple[str, date, date], ...],
    issue_label: str,
) -> None:
    """
    For each lane that still has ``since_cost``, remove it and ensure each finite split exists.
    Missing splits are **deep copies** of the former ``since`` row with updated ``Cost`` /
    ``Validity period``. Splits already present (e.g. from CSV) are left unchanged.
    """
    n = 0
    for ln in lanes:
        costs = ln.get("Costs") or []
        i_since = next(
            (
                i
                for i, c in enumerate(costs)
                if (c.get("Cost") or "").strip() == since_cost
            ),
            None,
        )
        if i_since is None:
            continue
        since_row = copy.deepcopy(costs[i_since])
        present_titles = {(c.get("Cost") or "").strip() for c in costs}
        new_costs = [c for i, c in enumerate(costs) if i != i_since]
        for cost_title, d_lo, d_hi in split_specs:
            if cost_title in present_titles:
                continue
            row = copy.deepcopy(since_row)
            row["Cost"] = cost_title
            row["Validity period"] = f"from {_fmt_dmy(d_lo)} to {_fmt_dmy(d_hi)}"
            new_costs.append(row)
        ln["Costs"] = new_costs
        n += 1
    if n:
        issues.append(
            f"  Card-wide *DFT* {issue_label}: migrated {n} lane(s) — removed {since_cost!r} and "
            f"ensured split columns {split_specs[0][0]!r} / {split_specs[1][0]!r} "
            f"(copied from former since row where missing)."
        )


def merge_rate_with_updates(
    rate: dict,
    merged_records: list[dict],
    update_label: str,
    issues: list[str],
) -> tuple[dict, set[tuple], set[tuple]]:
    """
    Returns (rate_copy, excel_green_cells, meta)
    excel_green_cells: set of (sheet, row_1b, col_1b) for Rate card

    Appends a step-by-step merge trace to ``issues`` (redundant-check and merge routing).
    """
    log.debug(
        "merge_rate_with_updates: %s combined CSV record(s); update_label=%r",
        len(merged_records),
        update_label,
    )
    rate = copy.deepcopy(rate)
    rc = rate["sheets"]["Rate card"]
    lanes = rc.get("lanes") or []
    rc["lanes"] = lanes
    cost_names_list: list[str] = list(rc.get("cost_names", []))
    cost_names_set = set(cost_names_list)

    if not lanes:
        issues.append(
            "Rate card has no lanes; donor templates are empty until lanes exist."
        )
        log.debug("merge: baseline Rate card has no lanes — donor lane templates empty.")

    donor = find_donor_lane(lanes, None)
    excel_marks: set[tuple] = set()
    log.debug(
        "merge: starting with %s lane(s) in baseline JSON; donor lane #%r for new cost scaffolding.",
        len(lanes),
        donor.get("Lane #"),
    )

    merged_records = sorted(merged_records, key=sort_key_combined_or_merge_record)
    nrec = len(merged_records)
    if nrec > 1:
        issues.append(
            "Merge applies combined rows in chronological order (first effective date first): "
            + "; ".join(
                f"{_fmt_dmy(sort_key_combined_or_merge_record(r)[0])}"
                f"–{_fmt_dmy(sort_key_combined_or_merge_record(r)[1])}"
                for r in merged_records
            )
        )

    for ri, rec in enumerate(merged_records):
        sub_recs = flat_records_from_combined_merge_rec(rec)
        base_rec = sub_recs[0]
        svc = service_cy_cy(base_rec.get("SERVICE__C", ""))
        origin = _norm(base_rec.get("ORIGIN_LOCATION_NAME__C"))
        dest = _norm(base_rec.get("DESTINATION_LOCATION_NAME__C"))
        car = _norm(base_rec.get("CARRIER"))

        ok_redundant, redundant_lines = merge_record_redundant_with_existing_lanes_explain(
            base_rec, lanes
        )
        win_note = (
            f" ({len(sub_recs)} fee windows)"
            if len(sub_recs) > 1
            else ""
        )
        issues.append(
            f"--- merge trace [{ri + 1}/{nrec}] KEY={base_rec.get('KEY')!r}{win_note} ---"
        )
        issues.extend(redundant_lines)
        if ok_redundant and not APPLY_CSV_UPDATES_EVEN_WHEN_REDUNDANT:
            issues.append(
                f"Skipped update for {base_rec.get('KEY')!r}: route lanes already have "
                f"matching tiered fees (price, currency, validity) for "
                f"{car} {origin}->{dest} ({svc})."
            )
            log.debug(
                "[%s/%s] REDUNDANT skip KEY=%r — existing lanes already satisfy tier/price.",
                ri + 1,
                nrec,
                base_rec.get("KEY"),
            )
            continue

        route_lanes_pre = lanes_matching_route(lanes, car, svc, origin, dest)
        if not route_lanes_pre and len(sub_recs) > 1:
            effs_b: list[date] = []
            exps_b: list[date] = []
            for rb in sub_recs:
                e = _parse_us_date(str(rb.get("RATE_EFFECTIVE_DATE__C") or "").strip())
                x = _parse_us_date(str(rb.get("RATE_EXPIRATION_DATE__C") or "").strip())
                if e is not None:
                    effs_b.append(e)
                if x is not None:
                    exps_b.append(x)
            eff_lane = min(effs_b) if effs_b else date.today()
            exp_lane = max(exps_b) if exps_b else date.today()
            combined_costs = build_lane_costs_for_new_route_batched(
                sub_recs, donor, lanes, issues
            )
            for c in combined_costs:
                nm = c.get("Cost")
                if nm and nm not in cost_names_set:
                    cost_names_list.append(nm)
                    cost_names_set.add(nm)
            nnums = next_lane_numbers(lanes, 1)
            ln = str(nnums[0])
            lanes.append(
                {
                    "Lane #": ln,
                    "KEY": base_rec.get("KEY"),
                    "Rate Card": update_label,
                    "Carrier": car,
                    "SERVICE": "not PRECARRIAGE/ONCARRIAGE",
                    "Service": svc,
                    "Valid from": _fmt_dmy(eff_lane),
                    "Valid to": _fmt_dmy(exp_lane),
                    "Origin Port": origin,
                    "Destination Port": dest,
                    "Costs": copy.deepcopy(combined_costs),
                }
            )
            issues.append(
                f"  merge-path: new route — single lane {ln} "
                f"{_fmt_dmy(eff_lane)}–{_fmt_dmy(exp_lane)} from {len(sub_recs)} "
                f"CSV fee window(s) on one combined row ({car} {origin}->{dest}, Service {svc})."
            )
            log.debug(
                "merge: batched %s fee window(s) for new route into lane %r",
                len(sub_recs),
                ln,
            )
            continue

        runs = (
            sub_recs
            if (route_lanes_pre and len(sub_recs) > 1)
            else [sub_recs[0]]
        )

        for win_idx, sr in enumerate(runs):
            if len(runs) > 1:
                issues.append(
                    f"  merge CSV window [{win_idx + 1}/{len(runs)}]: "
                    f"{_fmt_dmy(_parse_us_date(sr.get('RATE_EFFECTIVE_DATE__C')) or date.today())}"
                    f"–{_fmt_dmy(_parse_us_date(sr.get('RATE_EXPIRATION_DATE__C')) or date.today())}"
                )

            new_costs = build_lane_costs_from_update(sr, donor, lanes, issues)
            log.debug(
                "[%s/%s] built %s cost row(s) from CSV for KEY=%r",
                ri + 1,
                nrec,
                len(new_costs),
                sr.get("KEY"),
            )
            for c in new_costs:
                nm = c.get("Cost")
                if nm and nm not in cost_names_set:
                    cost_names_list.append(nm)
                    cost_names_set.add(nm)

            eff_s = _fmt_dmy(_parse_us_date(sr.get("RATE_EFFECTIVE_DATE__C")) or date.today())
            exp_s = _fmt_dmy(_parse_us_date(sr.get("RATE_EXPIRATION_DATE__C")) or date.today())

            route_lanes = lanes_matching_route(lanes, car, svc, origin, dest)
            upd_from_d = _parse_us_date(sr.get("RATE_EFFECTIVE_DATE__C"))
            upd_to_d = _parse_us_date(sr.get("RATE_EXPIRATION_DATE__C"))
            route_lanes_overlapping_csv: list[dict] = []
            if upd_from_d and upd_to_d:
                route_lanes_overlapping_csv = [
                    ln
                    for ln in route_lanes
                    if _lane_validity_overlaps_csv_window(ln, upd_from_d, upd_to_d)
                ]
                route_lanes_overlapping_csv.sort(key=_int_lane_num)
                issues.append(
                    f"  step {ri + 1}/{nrec}: CSV window "
                    f"{_fmt_dmy(upd_from_d)}–{_fmt_dmy(upd_to_d)} → overlapping lane(s): "
                    f"{[ln.get('Lane #') for ln in route_lanes_overlapping_csv]}"
                )
            # Primary lane for split / overlap checks must have CSV ∩ lane validity (not first route
            # lane by list order — older validity windows like ``01.06.2025`` would spuriously append).
            mlane = route_lanes_overlapping_csv[0] if route_lanes_overlapping_csv else None
            log.debug(
                "[%s/%s] route match: %s same-route lane(s), %s CSV-validity-overlap lane(s), primary=%r.",
                ri + 1,
                nrec,
                len(route_lanes),
                len(route_lanes_overlapping_csv),
                mlane.get("Lane #") if mlane else None,
            )
            prospect_names = {
                str(c.get("Cost")).strip() for c in new_costs if c.get("Cost")
            }
            lane_names: set[str] = set()
            for ln_r in route_lanes:
                for c in ln_r.get("Costs", []):
                    nm = c.get("Cost")
                    if nm:
                        lane_names.add(str(nm).strip())
            overlap = prospect_names & lane_names
            effective_overlap = bool(overlap)
    
            def costs_for_new_lane_append() -> list[dict]:
                return copy.deepcopy(new_costs)
    
            def append_fresh_lane() -> None:
                nnums = next_lane_numbers(lanes, 1)
                ln = str(nnums[0])
                lanes.append(
                    {
                        "Lane #": ln,
                        "KEY": sr.get("KEY"),
                        "Rate Card": update_label,
                        "Carrier": car,
                        "SERVICE": "not PRECARRIAGE/ONCARRIAGE",
                        "Service": svc,
                        "Valid from": eff_s,
                        "Valid to": exp_s,
                        "Origin Port": origin,
                        "Destination Port": dest,
                        "Costs": costs_for_new_lane_append(),
                    }
                )
                issues.append(
                    f"Appended new lane {ln} {car} {origin}->{dest} (Service {svc})."
                )

            # Spec (Ocean Rates 20260306): if the update KEY's equipment (e.g. 42UTH) is not
            # present on *any* matching-route lane, append a new lane — do not split or in-place
            # merge into a lane that only has other equipment (e.g. 45G0 on lane 4).
            eq_key = merge_routing_container_equipment(sr)
            if (
                mlane is not None
                and eq_key
                and rec_has_cntr_rate_column_values(sr)
                and not _route_has_equipment_on_any_lane(route_lanes, eq_key)
            ):
                issues.append(
                    "  merge-path: DECISION — equipment "
                    f"{eq_key!r} from KEY not on any cost row of this route → "
                    f"append_fresh_lane() (new equipment; do not alter existing lane validity)."
                )
                append_fresh_lane()
                continue
    
            if MERGE_FULL_TRACE:
                issues.append(
                    f"  merge-path: new_costs count={len(new_costs)} "
                    f"(from build_lane_costs_from_update)"
                )
                for c in new_costs[:8]:
                    issues.append(f"    → Cost={c.get('Cost')!r}")
                if len(new_costs) > 8:
                    issues.append(f"    → … ({len(new_costs) - 8} more)")
                rl_ids = [x.get("Lane #") for x in route_lanes[:12]]
                issues.append(
                    f"  merge-path: route lanes (Carrier/Service/ports) → "
                    f"{len(route_lanes)} lane(s): {rl_ids}"
                    + (" …" if len(route_lanes) > 12 else "")
                )
                issues.append(
                    f"  merge-path: primary lane for split/in-place → "
                    f"Lane #={repr(mlane.get('Lane #') if mlane else None)}"
                )
                issues.append(
                    f"  merge-path: cost name overlap (prospect ∩ all route lanes) = "
                    f"{len(overlap)} name(s): {sorted(overlap)[:10]}; "
                    f"effective_overlap={effective_overlap}"
                )
            else:
                issues.append(
                    f"  merge-path: costs={len(new_costs)} overlap={len(overlap)} "
                    f"effective={effective_overlap} lanes={len(route_lanes)} "
                    f"csv_overlap={ [x.get('Lane #') for x in route_lanes_overlapping_csv] } "
                    f"primary={mlane.get('Lane #') if mlane else None}"
                )
    
            if mlane is None:
                issues.append(
                    "  merge-path: DECISION — no lane matches route → append_fresh_lane()"
                )
                append_fresh_lane()
            elif (
                _row_only_dft_baf_eu_ets_rates_no_other_cost_columns(sr)
                and rec_has_dft_rate_column_values(sr)
                and route_lanes_overlapping_csv
            ):
                patch = filter_dft_fee_cost_rows_for_csv_window(new_costs, sr)
                if not patch and (
                    sr.get("*DFTBAF_rate (per w/m)") not in (None, "")
                    or sr.get("*DFTEU_ETS_rate (per w/m)") not in (None, "")
                ):
                    issues.append(
                        "  merge-path: warning — *DFT* window filter matched no tier rows "
                        f"(validity/price vs CSV {_fmt_dmy(upd_from_d)}–{_fmt_dmy(upd_to_d)}); "
                        "lanes unchanged for this window."
                    )
                else:
                    for ln_tgt in route_lanes_overlapping_csv:
                        patch_ln = filter_dft_fee_cost_rows_for_lane(patch, ln_tgt)
                        merge_dft_fee_cost_rows_into_lane(ln_tgt, patch_ln)
                    issues.append(
                        "  merge-path: DECISION — *DFT* BAF/EU ETS only: applied CSV window "
                        f"{_fmt_dmy(upd_from_d)}–{_fmt_dmy(upd_to_d)} to overlapping lane(s) "
                        f"{[x.get('Lane #') for x in route_lanes_overlapping_csv]} "
                        f"(≤{len(patch)} tier row(s) before lane validity trim; "
                        "historical donor tiers excluded)."
                    )
                continue
            elif not effective_overlap:
                issues.append(
                    f"  merge-path: DECISION — prospect_names ∩ lane_names is empty "
                    f"(CSV would add costs not named on lane {mlane.get('Lane #')}) "
                    f"→ append_fresh_lane() beside existing lane"
                )
                issues.append(
                    f"No cost names overlap with existing lane {mlane.get('Lane #')} "
                    f"{car} {origin}->{dest}; left it unchanged and appended a new lane."
                )
                append_fresh_lane()
            else:
                snap = copy.deepcopy(mlane)
                orig_from = _parse_dmy_text(_norm(snap.get("Valid from")))
                orig_to = _parse_dmy_text(_norm(snap.get("Valid to")))
                upd_from = _parse_us_date(sr.get("RATE_EFFECTIVE_DATE__C"))
                upd_to = _parse_us_date(sr.get("RATE_EXPIRATION_DATE__C"))

                split_handled = False
                use_fallback = False
                if not all([orig_from, orig_to, upd_from, upd_to]):
                    issues.append(
                        f"  merge-path: split not possible — missing dates "
                        f"(lane from/to or CSV): orig={orig_from} {orig_to} upd={upd_from} {upd_to}"
                    )
                    issues.append(
                        f"Split skipped for lane {mlane.get('Lane #')}: missing Valid from/to or CSV dates; "
                        "using in-place equipment replacement."
                    )
                    use_fallback = True
                elif upd_to < orig_from or upd_from > orig_to:
                    issues.append(
                        f"  merge-path: DECISION — CSV window does not overlap lane "
                        f"{mlane.get('Lane #')} validity "
                        f"({_fmt_dmy(orig_from)}–{_fmt_dmy(orig_to)}) → append_fresh_lane()"
                    )
                    issues.append(
                        f"Update window does not overlap lane {mlane.get('Lane #')} validity; "
                        "existing lane unchanged, appending new lane."
                    )
                    append_fresh_lane()
                    continue
                elif upd_from == orig_from and upd_to < orig_to:
                    # CSV window starts exactly when the lane starts but ends before the lane —
                    # keep the same Lane # through ``upd_to`` with merged costs, append one lane that
                    # duplicates pre-split costs for ``upd_to+1 .. orig_to`` (Ocean Rates *DFTBASE*
                    # scenario: first row trims lane 265; second row updates the continuation lane).
                    mlane["Valid to"] = _fmt_dmy(upd_to)
                    mlane["Costs"] = snapshot_costs_with_replaced_equipment(
                        snap.get("Costs", []), sr, new_costs
                    )
                    tail_ln = str(next_lane_numbers(lanes, 1)[0])
                    tail_from = upd_to + timedelta(days=1)
                    tail_lane = {
                        "Lane #": tail_ln,
                        "KEY": snap.get("KEY"),
                        "Rate Card": snap.get("Rate Card"),
                        "Carrier": _norm(snap.get("Carrier")),
                        "SERVICE": snap.get("SERVICE"),
                        "Service": snap.get("Service"),
                        "Valid from": _fmt_dmy(tail_from),
                        "Valid to": _fmt_dmy(orig_to),
                        "Origin Port": snap.get("Origin Port"),
                        "Destination Port": snap.get("Destination Port"),
                        "Costs": copy.deepcopy(snap.get("Costs", [])),
                    }
                    lanes.append(tail_lane)
                    issues.append(
                        "  merge-path: DECISION — lane-head validity split (CSV aligns with lane start): "
                        f"lane {snap.get('Lane #')} {_fmt_dmy(orig_from)}–{_fmt_dmy(upd_to)} merged; "
                        f"continuation lane {tail_ln} {_fmt_dmy(tail_from)}–{_fmt_dmy(orig_to)} "
                        "(duplicate costs for remainder)."
                    )
                    split_handled = True
                else:
                    pre_end = upd_from - timedelta(days=1)
                    if pre_end < orig_from:
                        issues.append(
                            "  merge-path: update starts strictly before lane start → "
                            "in-place replacement only (no split)"
                        )
                        issues.append(
                            f"Split skipped: update starts before lane start "
                            f"({_fmt_dmy(upd_from)} vs {_fmt_dmy(orig_from)}); in-place equipment replacement."
                        )
                        use_fallback = True
    
                if split_handled:
                    pass
                elif use_fallback:
                    mlane["Costs"] = snapshot_costs_with_replaced_equipment(
                        mlane.get("Costs", []), sr, new_costs
                    )
                    issues.append(
                        f"  merge-path: DECISION — in-place replace costs on lane {mlane.get('Lane #')}"
                    )
                    issues.append(
                        f"Updated existing lane {mlane.get('Lane #')} {car} {origin}->{dest} (Service {svc})."
                    )
                else:
                    issues.append(
                        f"  merge-path: DECISION — split lane {mlane.get('Lane #')}: "
                        f"pre [{_fmt_dmy(orig_from)}–{_fmt_dmy(pre_end)}], "
                        f"middle [{_fmt_dmy(upd_from)}–{_fmt_dmy(upd_to)}], "
                        f"tail from {_fmt_dmy(upd_to + timedelta(days=1))}"
                    )
                    mlane["Valid to"] = _fmt_dmy(pre_end)
                    mid_ln = str(next_lane_numbers(lanes, 1)[0])
                    middle_lane = {
                        "Lane #": mid_ln,
                        "KEY": sr.get("KEY"),
                        "Rate Card": update_label,
                        "Carrier": car,
                        "SERVICE": _norm(snap.get("SERVICE"))
                        or "not PRECARRIAGE/ONCARRIAGE",
                        "Service": svc,
                        "Valid from": _fmt_dmy(upd_from),
                        "Valid to": _fmt_dmy(upd_to),
                        "Origin Port": origin,
                        "Destination Port": dest,
                        "Costs": snapshot_costs_with_replaced_equipment(
                            snap.get("Costs", []), sr, new_costs
                        ),
                    }
                    lanes.append(middle_lane)
    
                    tail_from = upd_to + timedelta(days=1)
                    if tail_from <= orig_to:
                        tail_ln = str(next_lane_numbers(lanes, 1)[0])
                        tail_lane = {
                            "Lane #": tail_ln,
                            "KEY": snap.get("KEY"),
                            "Rate Card": snap.get("Rate Card"),
                            "Carrier": _norm(snap.get("Carrier")),
                            "SERVICE": snap.get("SERVICE"),
                            "Service": snap.get("Service"),
                            "Valid from": _fmt_dmy(tail_from),
                            "Valid to": _fmt_dmy(orig_to),
                            "Origin Port": origin,
                            "Destination Port": dest,
                            "Costs": copy.deepcopy(snap.get("Costs", [])),
                        }
                        lanes.append(tail_lane)
                        issues.append(
                            f"Split existing lane {snap.get('Lane #')}: pre-period "
                            f"{_fmt_dmy(orig_from)}–{_fmt_dmy(pre_end)}; update lane {mid_ln} "
                            f"{_fmt_dmy(upd_from)}–{_fmt_dmy(upd_to)}; continuation {tail_ln} "
                            f"{_fmt_dmy(tail_from)}–{_fmt_dmy(orig_to)}."
                        )
                    else:
                        issues.append(
                            f"Split existing lane {snap.get('Lane #')}: pre-period "
                            f"{_fmt_dmy(orig_from)}–{_fmt_dmy(pre_end)}; update lane {mid_ln} "
                            f"{_fmt_dmy(upd_from)}–{_fmt_dmy(upd_to)}; no tail (update reaches lane end)."
                        )

    log.debug(
        "merge: pre-consolidation lane count=%s — running consolidate_mergeable_route_lanes.",
        len(lanes),
    )
    lanes = consolidate_mergeable_route_lanes(lanes, issues)
    for ln in lanes:
        ln["Costs"] = _dedupe_duplicate_cost_names_on_lane(ln.get("Costs", []))
    run_baf_since_split = (
        not MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES
        or merged_records_touch_dft_baf_fee(merged_records)
    )
    run_eu_ets_since_split = (
        not MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES
        or merged_records_touch_dft_eu_ets_fee(merged_records)
    )
    if run_baf_since_split:
        _migrate_since_tier_to_split_columns(
            lanes,
            issues,
            since_cost=BAF_SINCE_2026_COST,
            split_specs=BAF_SPLIT_WINDOWS_2026,
            issue_label="BAF",
        )
    elif MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES:
        issues.append(
            "  post-merge: skipped BAF ``since``→split migration — CSV has no *DFTBAF_* fee values."
        )
    if run_eu_ets_since_split:
        _migrate_since_tier_to_split_columns(
            lanes,
            issues,
            since_cost=EU_ETS_SINCE_2026_COST,
            split_specs=EU_ETS_SPLIT_WINDOWS_2026,
            issue_label="EU ETS",
        )
    elif MIGRATE_SINCE_TIER_SPLITS_ONLY_WHEN_CSV_HAS_DFT_FEES:
        issues.append(
            "  post-merge: skipped EU ETS ``since``→split migration — CSV has no *DFTEU_ETS_* fee values."
        )
    for ln in lanes:
        prune_dft_breakbulk_fees_outside_lane_window(ln)
    for ln in lanes:
        ln["Costs"] = _dedupe_duplicate_cost_names_on_lane(ln.get("Costs", []))
    rc["lanes"] = lanes
    log.debug("merge: post-consolidation lane count=%s", len(lanes))

    merged_cost_names: list[str] = []
    seen_cn: set[str] = set()
    for nm in cost_names_list:
        if nm and str(nm).strip():
            s = str(nm).strip()
            if s not in seen_cn:
                seen_cn.add(s)
                merged_cost_names.append(s)
    for ln in lanes:
        for c in ln.get("Costs", []):
            nm = c.get("Cost")
            if nm and str(nm).strip():
                s = str(nm).strip()
                if s not in seen_cn:
                    seen_cn.add(s)
                    merged_cost_names.append(s)
    rc["cost_names"] = sort_rate_card_cost_names(merged_cost_names)
    for ln in lanes:
        _sort_lane_costs_in_column_order(ln)
    log.debug(
        "merge_rate_with_updates: finished — %s lane(s), %s distinct cost name(s) in cost_names.",
        len(lanes),
        len(rc["cost_names"]),
    )
    return rate, excel_marks, set()


def find_cost_entry(rate: dict, cost_name: str) -> dict | None:
    for ln in rate["sheets"]["Rate card"]["lanes"]:
        for c in ln.get("Costs", []):
            if c.get("Cost") == cost_name:
                return c
    return None


def _lane_cost_map(ln: dict) -> dict[str, tuple[str, object]]:
    """``Cost`` name → (currency, price) for merge compatibility checks."""
    out: dict[str, tuple[str, object]] = {}
    for c in ln.get("Costs") or []:
        nm = c.get("Cost")
        if not nm:
            continue
        out[str(nm).strip()] = (_norm(c.get("Currency")), c.get("Price"))
    return out


def _cost_maps_mergeable(m1: dict, m2: dict) -> bool:
    """True if shared cost names agree on currency and price."""
    for k in set(m1) & set(m2):
        if m1[k] != m2[k]:
            return False
    return True


def _parse_leading_date_range_in_parens(
    s: str,
) -> tuple[date | None, date | None] | None:
    """
    First ``(dd.mm.yyyy-dd.mm.yyyy)`` in the string, or
    ``(dd.mm.yyyy - dd.mm.yyyy)`` (spaced hyphen).

    Also ``(45G0 15.01.2026-31.05.2026)`` / ``(42UTH 25.02.2026 - 31.03.2026)`` where the
    range follows container equipment in the same parentheses.
    """
    m = re.search(
        r"\((\d{1,2}\.\d{1,2}\.\d{4})\s*[-–]\s*(\d{1,2}\.\d{1,2}\.\d{4})\)",
        s,
    )
    if m:
        a, b = _parse_dmy_text(m.group(1)), _parse_dmy_text(m.group(2))
        if a and b:
            return (a, b)
    m2 = re.search(
        r"\([0-9]{2}[A-Za-z0-9]+\s+(\d{1,2}\.\d{1,2}\.\d{4})\s*[-–]\s*(\d{1,2}\.\d{1,2}\.\d{4})\)",
        s,
    )
    if m2:
        a, b = _parse_dmy_text(m2.group(1)), _parse_dmy_text(m2.group(2))
        if a and b:
            return (a, b)
    return None


def _parse_since_date_in_title(s: str) -> date | None:
    m = re.search(r"since\s+(\d{1,2}\.\d{1,2}\.\d{4})", s, re.I)
    if not m:
        return None
    return _parse_dmy_text(m.group(1))


def _extract_container_equipment_code(cost_name: str) -> str | None:
    """First container-style code in parentheses, e.g. ``45G0`` in ``Base Rate (45G0)``."""
    m = re.match(
        r"^(?:BAF Fee|Base Rate|EU ETS Fee|Destination Terminal Handling Fee|"
        r"Origin Terminal Handling Fee|Dangerous Goods Fee)\s*\(([0-9]{2}[A-Za-z0-9]+)",
        cost_name.strip(),
    )
    return m.group(1) if m else None


def _cost_name_inner_tuple(s: str) -> tuple:
    """
    Ordering **within** one equipment group: BAF → Base → Dangerous → DTHC → OTHC → EU ETS,
    with dated BAF/ETS chronology where applicable (matches ocean column layout spec).
    """
    sl = s.casefold()
    far = (999, 9, 99999999, 99999999, 9, s.casefold())

    if sl.startswith("baf fee"):
        if "(" not in s:
            return (5, 9, 0, 0, 0, 0, s.casefold())
        pr = _parse_leading_date_range_in_parens(s)
        if pr and pr[0] and pr[1]:
            a, b = pr
            return (5, 0, a.toordinal(), b.toordinal(), 0, 0, s.casefold())
        sd = _parse_since_date_in_title(s)
        if sd:
            return (5, 1, sd.toordinal(), 99999999, 0, 0, s.casefold())
        return (5, 2, 0, 0, 0, 0, s.casefold())

    if sl.startswith("base rate"):
        if "(BASE_" in s or "DFTROLL_" in s:
            return (10, 0, 0, 0, 0, 0, s.casefold())
        return (11, 0, 0, 0, 0, 0, s.casefold())
    if "dangerous goods" in sl:
        return (15, 0, 0, 0, 0, 0, s.casefold())
    if "destination terminal" in sl and "fee" in sl:
        return (20, 0, 0, 0, 0, 0, s.casefold())
    if "origin terminal" in sl and "fee" in sl:
        return (25, 0, 0, 0, 0, 0, s.casefold())

    if sl.startswith("eu ets fee"):
        if sl == "eu ets fee" or s == "EU ETS Fee":
            return (30, 6, 99999999, 99999999, 9, 0, s.casefold())
        if _is_container_style_eu_ets(s):
            return (30, 1, 0, 0, 0, 0, s.casefold())
        pr = _parse_leading_date_range_in_parens(s)
        if pr and pr[0] and pr[1]:
            a, b = pr
            return (30, 3, a.toordinal(), b.toordinal(), 0, 0, s.casefold())
        sd = _parse_since_date_in_title(s)
        if sd:
            return (30, 4, sd.toordinal(), 99999999, 0, 0, s.casefold())
        return (30, 5, 0, 0, 0, 0, s.casefold())

    return far


def _sort_key_cost_column_name(name: str) -> tuple:
    """
    Primary: **equipment** (45G0, 42UTH, …) so all fees for one equipment stay adjacent.
    Secondary: fee type / validity (see ``_cost_name_inner_tuple``).
    *DFT* breakbulk ``BASE_*`` / ``DFTROLL_*`` columns sort after numbered equipment using ``EQ_SORT_ORDER`` spare bucket.
    """
    s = str(name).strip()
    eq = _extract_container_equipment_code(s)
    if eq is not None:
        grp = EQ_SORT_ORDER.get(eq, 55)
    elif "(BASE_" in s or "DFTROLL_" in s or ("Base Rate (BASE_" in s):
        grp = 72
    elif _is_breakbulk_non_container_cost_name(s):
        grp = 73
    else:
        grp = 71
    inner = _cost_name_inner_tuple(s)
    return (grp, eq or "", inner, s.casefold())


def _is_breakbulk_non_container_cost_name(s: str) -> bool:
    """*DFT* BAF / EU ETS without container equipment digits."""
    if s.startswith("EU ETS Fee") and not _is_container_style_eu_ets(s):
        return True
    if s.startswith("BAF Fee") and not _is_container_style_baf(s):
        return True
    return False


def sort_rate_card_cost_names(names: list[str]) -> list[str]:
    return sorted(names, key=_sort_key_cost_column_name)


def _sort_lane_costs_in_column_order(ln: dict) -> None:
    costs = ln.get("Costs") or []
    with_name = [c for c in costs if c.get("Cost")]
    without = [c for c in costs if not c.get("Cost")]
    if not with_name:
        return
    by_name: dict[str, dict] = {}
    for c in with_name:
        by_name[str(c.get("Cost")).strip()] = c
    order = sort_rate_card_cost_names(list(by_name.keys()))
    ln["Costs"] = [by_name[k] for k in order] + without


def _last_used_row_col_a(ws, header_row: int) -> int:
    """Last row at or below header that has a non-empty column A (avoids trailing blank rows)."""
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def cost_needs_min_layout(cost_entry: dict | None) -> bool:
    """True when JSON carries a non-zero Minimum (Flat column required in Excel)."""
    if not cost_entry:
        return False
    m = cost_entry.get("Minimum")
    if m is None or m == "":
        return False
    try:
        return float(m) != 0.0
    except (TypeError, ValueError):
        return True


def infer_excel_cost_block_width(ws, header_row: int, anchor_col: int) -> int:
    """
    Detect existing **lane-header row** layout from the sheet alone:

    - **3** — Currency | MIN/Flat | p/unit (middle cell often ``MIN\\nFlat``).
    - **2** — Currency | p/unit only.

    Cost blocks with JSON ``Minimum`` use **three** underlying columns (title row merged
    across three cells); usual costs use **two**. Prefer :func:`resolve_excel_cost_block_width`
    when a merged JSON cost row is available.
    """
    mid = ws.cell(row=header_row, column=anchor_col + 1).value
    right = ws.cell(row=header_row, column=anchor_col + 2).value
    mid_s = str(mid).strip().lower() if mid is not None else ""
    right_s = str(right).strip().lower() if right is not None else ""
    if ("flat" in mid_s or "min" in mid_s) and (
        right_s.startswith("p") or "unit" in right_s
    ):
        return 3
    return 2


def _title_row_merge_span_width(ws, cost_row: int, anchor_col: int) -> int | None:
    """
    If the fee **title** row has a horizontal merge containing ``anchor_col``, return its
    width in columns (**2** or **3**). Matches Transporeon: MIN costs merge the title across
    three columns; usual costs across two.
    """
    for mr in list(ws.merged_cells):
        if mr.min_row != mr.max_row:
            continue
        if mr.min_row != cost_row:
            continue
        if not (mr.min_col <= anchor_col <= mr.max_col):
            continue
        w = mr.max_col - mr.min_col + 1
        if w in (2, 3):
            return w
    return None


def resolve_excel_cost_block_width(
    ws,
    cost_row: int,
    header_row: int,
    anchor_col: int,
    cost_entry: dict | None,
) -> int:
    """
    Column span for one cost: **3** when JSON has non-zero ``Minimum`` *or* the sheet
    already shows a 3-column MIN block / 3-wide title merge; **2** for a normal fee.

    Use this when writing headers or lane cells so MIN lanes match three merged columns.
    """
    if cost_entry is not None:
        if cost_needs_min_layout(cost_entry):
            return 3
        if infer_excel_cost_block_width(ws, header_row, anchor_col) >= 3:
            return 3
        tw = _title_row_merge_span_width(ws, cost_row, anchor_col)
        if tw == 3:
            return 3
        return 2
    tw = _title_row_merge_span_width(ws, cost_row, anchor_col)
    if tw in (2, 3):
        return tw
    return infer_excel_cost_block_width(ws, header_row, anchor_col)


def _force_unmerge_range(ws, mr) -> bool:
    """
    Remove a merged range and drop ``MergedCell`` placeholders for non-top-left cells.

    ``merged_cells`` may store a ``CellRange`` not identical (``is``) to ``mr`` from a
    stale iterator, so we retry removal by bounds. ``Worksheet.unmerge_cells`` is not
    used here (coord / subset quirks).
    """
    resolved: object | None = None
    by_bounds = CellRange(
        min_row=mr.min_row,
        min_col=mr.min_col,
        max_row=mr.max_row,
        max_col=mr.max_col,
    )
    for cand in (mr, by_bounds):
        try:
            ws.merged_cells.remove(cand)
            resolved = cand
            break
        except KeyError:
            continue
    if resolved is None:
        b = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
        for existing in list(ws.merged_cells):
            if (
                existing.min_row,
                existing.min_col,
                existing.max_row,
                existing.max_col,
            ) == b:
                try:
                    ws.merged_cells.remove(existing)
                    resolved = existing
                    break
                except KeyError:
                    continue
    if resolved is None:
        return False
    cells = list(resolved.cells)
    for row, col in cells[1:]:
        ws._cells.pop((row, col), None)
    return True


def _ensure_writable_cell(ws, row: int, col: int) -> None:
    """If ``_cells`` still holds a ``MergedCell`` at ``(row,col)``, replace with ``Cell``."""
    coord = (row, col)
    if isinstance(ws._cells.get(coord), MergedCell):
        ws._cells.pop(coord, None)
        ws._cells[coord] = Cell(ws, row=row, column=col)


def _unmerge_cells_overlapping(
    ws, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    """
    Unmerge any merged range that touches the rectangle. Repeats because unmerging can
    split ranges; using exact range objects avoids openpyxl ``unmerge_cells`` false
    ``ValueError: ... is not merged`` for equivalent string coordinates.
    """
    for _ in range(32):
        removed_any = False
        for mr in list(ws.merged_cells):
            if (
                mr.min_row <= max_row
                and mr.max_row >= min_row
                and mr.min_col <= max_col
                and mr.max_col >= min_col
            ):
                if _force_unmerge_range(ws, mr):
                    removed_any = True
        if not removed_any:
            break


def _unmerge_merged_cells_covering(ws, row: int, col: int) -> None:
    """
    Remove any merge range that contains ``(row, col)`` until none remain.
    """
    for _ in range(64):
        removed = False
        for mr in list(ws.merged_cells):
            if (
                mr.min_row <= row <= mr.max_row
                and mr.min_col <= col <= mr.max_col
            ):
                if _force_unmerge_range(ws, mr):
                    removed = True
        if not removed:
            break


def _merge_row_span(ws, row: int, min_col: int, max_col: int) -> None:
    if min_col >= max_col:
        return
    _unmerge_cells_overlapping(ws, row, row, min_col, max_col)
    ws.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)


def _copy_cell_visual_style(src: Cell, dst: Cell) -> None:
    """Copy font, fill, border, alignment, number format (no protection)."""
    if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
        return
    try:
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if getattr(src, "number_format", None):
            dst.number_format = src.number_format
    except (AttributeError, TypeError, ValueError):
        pass


def _style_reference_anchor_col(
    ws, cost_row: int, cost_name_to_col: dict[str, int]
) -> int:
    """
    Prefer a cost column whose fee title is still present (template look); else the
    leftmost mapped column.
    """
    if not cost_name_to_col:
        return 11
    for _name, ac in sorted(cost_name_to_col.items(), key=lambda x: x[1]):
        tit = merged_cell_top_left_value(ws, cost_row, ac)
        if tit is not None and str(tit).strip():
            return ac
    return min(cost_name_to_col.values())


def _apply_restored_title_band_style(
    ws,
    style_ref_col: int,
    cost_row: int,
    anchor: int,
) -> None:
    """Match title / validity / rate-by / conditional rows to the reference cost block."""
    for r in range(cost_row, cost_row + 4):
        _copy_cell_visual_style(ws.cell(r, style_ref_col), ws.cell(r, anchor))


def _fee_split_clone_source_col(
    new_name: str,
    cost_name_to_col: dict[str, int],
) -> int | None:
    """
    For a new **range-titled** fee (e.g. ``BAF Fee (15.01.2026 - 14.04.2026)``), return the
    template column of the donor ``… (since …)`` block to copy **Excel** title-band styling
    from (font/fill/border on the fee header rows).
    """
    name = str(new_name).strip()
    if "since" in name.lower():
        return None
    if "(" not in name or " - " not in name:
        return None
    prefix = name.split("(", 1)[0].strip()
    if not prefix:
        return None
    for nm, col in cost_name_to_col.items():
        sn = str(nm).strip()
        if not sn.startswith(prefix):
            continue
        if "since" in sn.lower():
            return col
    return None


def _clone_title_band_style_from_donor_block(
    ws,
    cost_row: int,
    donor_anchor: int,
    dest_anchor: int,
    layout_cols: int,
) -> None:
    """Copy visual style for the fee **title band** (4 rows) across a 2- or 3-col block."""
    for r in range(cost_row, cost_row + 4):
        for k in range(layout_cols):
            _copy_cell_visual_style(
                ws.cell(r, donor_anchor + k),
                ws.cell(r, dest_anchor + k),
            )


def _clear_excel_lane_row_unused_fee_columns(
    ws,
    lane_row: int,
    header_row: int,
    cost_row: int,
    cost_name_to_col: dict[str, int],
    lane_cost_names: set[str],
    rate: dict,
) -> None:
    """
    Blank Currency / MIN / p/unit lane cells for fee columns **not** used by this lane.
    Prevents legacy template values (e.g. ``BAF Fee (since …)``) from appearing on new lanes
    that only use split columns from JSON.
    """
    for cost_name, anchor in cost_name_to_col.items():
        if cost_name in lane_cost_names:
            continue
        ce = find_cost_entry(rate, cost_name)
        if not ce:
            ce = {}
        w_blk = resolve_excel_cost_block_width(ws, cost_row, header_row, anchor, ce)
        if w_blk == 3:
            ws.cell(lane_row, anchor).value = None
            ws.cell(lane_row, anchor + 1).value = None
            ws.cell(lane_row, anchor + 2).value = None
        else:
            ws.cell(lane_row, anchor).value = None
            ws.cell(lane_row, anchor + 1).value = None


def _write_excel_cost_text_block(
    ws,
    cost_row: int,
    c: int,
    cost_entry: dict,
    *,
    write_name: bool = True,
    layout_cols: int = 2,
) -> bool:
    """
    Write cost title row + validity / Apply if + Rate by + Rule (rows ``cost_row`` … +2).
    ``layout_cols`` 3 merges text across Currency | Flat | p/unit (MIN value goes on lane rows, not here).
    Returns True if ``missing_cost_template`` (Rate by intentionally blank — use red fill).
    """
    layout_cols = 3 if layout_cols == 3 else 2
    last_c = c + layout_cols - 1
    # Always clear merges touching this block (2- or 3-col); else anchor cells stay MergedCell.
    _unmerge_cells_overlapping(ws, cost_row, cost_row + 3, c, last_c)
    for rr, cc in (
        (cost_row, c),
        (cost_row + 1, c),
        (cost_row + 2, c),
    ):
        _unmerge_merged_cells_covering(ws, rr, cc)
        _ensure_writable_cell(ws, rr, cc)
    if write_name:
        ws.cell(cost_row, c).value = cost_entry.get("Cost") or ""
        if layout_cols == 3:
            _merge_row_span(ws, cost_row, c, last_c)
    v = cost_entry.get("Validity period") or ""
    cp = cost_entry.get("Cost to prolong") or ""
    ap = cost_entry.get("Apply if") or ""
    mn = cost_entry.get("Minimum")
    chunks = []
    if str(v).strip():
        chunks.append(f"Validity period: {v}")
    if str(cp).strip():
        chunks.append(f"Cost to prolong: {cp}")
    # With Flat/MIN columns, do not duplicate Minimum text in the validity block.
    if layout_cols == 2 and mn not in (None, ""):
        chunks.append(f"Minimum: {mn}")
    if str(ap).strip():
        chunks.append(ap)
    txt = "\n\n".join(chunks) if chunks else None
    ws.cell(cost_row + 1, c).value = txt
    if layout_cols == 3:
        _merge_row_span(ws, cost_row + 1, c, last_c)
    rb = cost_entry.get("Rate by") or ""
    ru = cost_entry.get("Rule") or ""
    if str(rb).strip() or str(ru).strip():
        ws.cell(cost_row + 2, c).value = (
            f"Rate by: {rb}\n{ru}" if rb and ru else (rb or ru)
        )
    else:
        ws.cell(cost_row + 2, c).value = None
    if layout_cols == 3:
        _merge_row_span(ws, cost_row + 2, c, last_c)
    return bool(cost_entry.get("missing_cost_template"))


def append_excel_cost_block(
    ws,
    cost_row: int,
    header_row: int,
    start_col: int,
    cost_entry: dict,
    ref_col: int,
    mark_green,
    mark_red,
    conditional_template=None,
    *,
    layout_cols: int = 2,
) -> None:
    layout_cols = 3 if layout_cols == 3 else 2
    last_c = start_col + layout_cols - 1
    miss = _write_excel_cost_text_block(
        ws,
        cost_row,
        start_col,
        cost_entry,
        write_name=True,
        layout_cols=layout_cols,
    )
    ct_val = (
        conditional_template
        if conditional_template is not None
        else ws.cell(cost_row + 3, ref_col).value
    )
    # Template merges make non-anchor cells read-only — clear merges before writing.
    _unmerge_cells_overlapping(ws, cost_row + 3, cost_row + 3, start_col, last_c)
    ws.cell(cost_row + 3, start_col).value = ct_val
    if layout_cols >= 2:
        _merge_row_span(ws, cost_row + 3, start_col, last_c)

    _unmerge_cells_overlapping(ws, header_row - 1, header_row, start_col, last_c)
    if layout_cols == 3:
        # ``header_row - 1`` is the conditional template row (same as ``cost_row + 3``); do not use it
        # for MIN — put **MIN** above **Flat** in the Lane-header row like Transporeon (stacked label).
        ws.cell(header_row, start_col).value = "Currency"
        ws.cell(header_row, start_col + 1).value = "MIN\nFlat"
        ws.cell(header_row, start_col + 2).value = "p/unit"
        mark_cols = list(range(start_col, start_col + 3))
    else:
        ws.cell(header_row - 1, start_col).value = None
        ws.cell(header_row - 1, start_col + 1).value = None
        ws.cell(header_row, start_col).value = "Currency"
        ws.cell(header_row, start_col + 1).value = "p/unit"
        mark_cols = [start_col, start_col + 1]

    for rr in range(cost_row, header_row + 1):
        for cc in mark_cols:
            if miss and rr == cost_row + 2 and cc == start_col:
                mark_red(rr, cc)
            else:
                mark_green(rr, cc)


def _next_free_cost_anchor_col(
    ws,
    cost_row: int,
    header_row: int,
    cost_name_to_col: dict[str, int],
    rate: dict,
) -> int:
    """First column after the rightmost cost block (2- vs 3-col from JSON MIN + sheet)."""
    end = 10
    for name, anchor in cost_name_to_col.items():
        ce = find_cost_entry(rate, name)
        w = resolve_excel_cost_block_width(ws, cost_row, header_row, anchor, ce)
        end = max(end, anchor + w - 1)
    return max(end + 1, 11)


def insert_excel_cost_columns_alphabetically(
    ws,
    cost_row: int,
    header_row: int,
    cost_name_to_col: dict[str, int],
    missing_names: list[str],
    rate: dict,
    ref_col: int,
    report: list[str],
    mark_green,
    mark_red,
    conditional_template=None,
) -> None:
    """
    Insert each missing cost as **2** columns (Currency | p/unit) or **3** (MIN layout) at the
    correct fee-order boundary. Inserts run **right-to-left** (highest ``insert_before`` first)
    so columns to the **left** of each insertion stay at the same indices until that insert runs —
    existing merged blocks are shifted as whole columns, not split. When two new costs share the
    same boundary, the **leftmost** fee in sort order is inserted first at that boundary.
    """
    log.debug(
        "insert_excel_cost_columns_alphabetically: ref_col=%s missing=%s (insert_before from fee sort vs anchors)",
        ref_col,
        missing_names[:40],
    )
    remaining = {
        str(n).strip()
        for n in missing_names
        if n and str(n).strip() not in cost_name_to_col
    }

    def _insert_before_for_new_name(new_name: str) -> int:
        kn = _sort_key_cost_column_name(new_name)
        for en in sorted(cost_name_to_col.keys(), key=_sort_key_cost_column_name):
            if kn < _sort_key_cost_column_name(en):
                return cost_name_to_col[en]
        return _next_free_cost_anchor_col(
            ws, cost_row, header_row, cost_name_to_col, rate
        )

    while remaining:
        candidates: list[tuple] = []
        for new_name in sorted(remaining, key=_sort_key_cost_column_name):
            ce = find_cost_entry(rate, new_name)
            if not ce:
                report.append(f"Skip Excel column for {new_name!r}: no JSON entry.")
                remaining.discard(new_name)
                continue
            ib = _insert_before_for_new_name(new_name)
            sk = _sort_key_cost_column_name(new_name)
            layout_cols = 3 if cost_needs_min_layout(ce) else 2
            candidates.append((ib, sk, new_name, ce, layout_cols))
        if not candidates:
            break
        # Rightmost insertion index first; tie → smaller fee key (leftmost new block at same pivot).
        candidates.sort(key=lambda t: (-t[0], t[1]))
        insert_before_col, _sk, new_name, ce, layout_cols = candidates[0]
        remaining.remove(new_name)

        donor_before = _fee_split_clone_source_col(new_name, cost_name_to_col)

        ws.insert_cols(insert_before_col, layout_cols)
        for k in list(cost_name_to_col.keys()):
            if cost_name_to_col[k] >= insert_before_col:
                cost_name_to_col[k] += layout_cols
        if donor_before is not None:
            if donor_before >= insert_before_col:
                donor_after = donor_before + layout_cols
            else:
                donor_after = donor_before
        else:
            donor_after = None

        if donor_after is not None:
            _clone_title_band_style_from_donor_block(
                ws, cost_row, donor_after, insert_before_col, layout_cols
            )

        append_excel_cost_block(
            ws,
            cost_row,
            header_row,
            insert_before_col,
            ce,
            ref_col,
            mark_green,
            mark_red,
            conditional_template,
            layout_cols=layout_cols,
        )

        cost_name_to_col[new_name] = insert_before_col
        report.append(
            f"Inserted Excel columns ({layout_cols}-wide) for cost {new_name!r} at col {insert_before_col}."
        )


def expand_template_cost_blocks_for_minimum(
    ws,
    cost_row: int,
    header_row: int,
    cost_name_to_col: dict[str, int],
    rate: dict,
    report: list[str],
    mark_green,
    mark_red,
    ref_col: int,
    conditional_template,
) -> None:
    """
    When merged JSON has non-zero ``Minimum`` but the template only has Currency | p/unit,
    insert a **Flat** column (MIN label above) between Currency and rate.
    Process **right-to-left** on anchors so column indices stay stable.
    """
    first: dict[str, dict] = {}
    for ln in rate.get("sheets", {}).get("Rate card", {}).get("lanes", []):
        for ce in ln.get("Costs", []):
            nm = ce.get("Cost")
            if nm:
                s = str(nm).strip()
                if s not in first:
                    first[s] = ce
    for name, anchor in sorted(cost_name_to_col.items(), key=lambda x: -x[1]):
        ce = first.get(name)
        if not ce or not cost_needs_min_layout(ce):
            continue
        if infer_excel_cost_block_width(ws, header_row, anchor) >= 3:
            continue
        insert_at = anchor + 1
        ws.insert_cols(insert_at, 1)
        for k in list(cost_name_to_col.keys()):
            if cost_name_to_col[k] >= insert_at:
                cost_name_to_col[k] += 1
        _unmerge_cells_overlapping(ws, cost_row, header_row, anchor, anchor + 2)
        ws.cell(header_row, anchor).value = "Currency"
        ws.cell(header_row, anchor + 1).value = "MIN\nFlat"
        ws.cell(header_row, anchor + 2).value = "p/unit"
        _write_excel_cost_text_block(
            ws, cost_row, anchor, ce, write_name=True, layout_cols=3
        )
        ct_val = (
            conditional_template
            if conditional_template is not None
            else ws.cell(cost_row + 3, ref_col).value
        )
        ws.cell(cost_row + 3, anchor).value = ct_val
        _merge_row_span(ws, cost_row + 3, anchor, anchor + 2)
        for rr in range(cost_row, header_row + 1):
            for cc in (anchor, anchor + 1, anchor + 2):
                mark_green(rr, cc)
        report.append(
            f"Expanded Excel block for {name!r} to MIN layout (Currency | Flat | p/unit) at col {anchor}."
        )


def _lane_currency_anchor_columns(ws, header_row: int, min_col: int = 11) -> list[int]:
    """
    Column indices where the lane-header row shows **Currency** — left edge of each fee’s
    Currency | (MIN) | p/unit band. Uses merge-aware reads so merged cells still resolve.
    """
    max_c = max(int(ws.max_column or 0), effective_max_column(ws))
    anchors: list[int] = []
    for c in range(min_col, max_c + 1):
        v = merged_cell_top_left_value(ws, header_row, c)
        if v is not None and str(v).strip().casefold() == "currency":
            anchors.append(c)
    return anchors


def _cost_title_row_merge_left_edges(
    ws, cost_row: int, min_col: int = 11, max_col: int | None = None
) -> list[int]:
    """
    Left columns of **single-row** horizontal merges on the fee **title** row. Used when
    the lane header has no **Currency** labels but title-row merges still mark each block.
    """
    if max_col is None:
        max_col = max(int(ws.max_column or 0), effective_max_column(ws))
    edges: set[int] = set()
    for mr in list(ws.merged_cells):
        try:
            if mr.min_row != mr.max_row or mr.min_row != cost_row:
                continue
            if mr.min_col < min_col or mr.min_col > max_col:
                continue
            w = mr.max_col - mr.min_col + 1
            if w < 1 or w > 4:
                continue
        except (AttributeError, TypeError, ValueError):
            continue
        edges.add(mr.min_col)
    return sorted(edges)


def _structural_cost_block_anchors(
    ws, cost_row: int, header_row: int, min_col: int = 11
) -> list[int]:
    """Prefer Currency-row anchors; otherwise title-row merge left edges."""
    ac = _lane_currency_anchor_columns(ws, header_row, min_col)
    if ac:
        return ac
    return _cost_title_row_merge_left_edges(ws, cost_row, min_col)


def restore_unmapped_cost_title_bands_from_rate(
    ws,
    cost_row: int,
    header_row: int,
    cost_name_to_col: dict[str, int],
    rate: dict,
    report: list[str],
) -> None:
    """
    Fill fee **title / validity / Rate by / conditional** rows from merged JSON when the
    Excel cells exist but text was lost (empty merged titles were skipped in the initial
    name→col scan). Optionally binds **unmapped** JSON cost names to **free** structural
    columns (Currency anchors or title-merge edges). Does **not** apply green or red fill —
    this is restoration, not inserting new structure.
    """
    ref_col = min(cost_name_to_col.values()) if cost_name_to_col else 11
    ct_val = ws.cell(cost_row + 3, ref_col).value
    style_ref_col = _style_reference_anchor_col(ws, cost_row, cost_name_to_col)

    def write_band(anchor: int, ce: dict) -> None:
        w = resolve_excel_cost_block_width(ws, cost_row, header_row, anchor, ce)
        last_c = anchor + w - 1
        _write_excel_cost_text_block(
            ws, cost_row, anchor, ce, write_name=True, layout_cols=w
        )
        _unmerge_cells_overlapping(ws, cost_row + 3, cost_row + 3, anchor, last_c)
        ws.cell(cost_row + 3, anchor).value = ct_val
        if last_c > anchor:
            _merge_row_span(ws, cost_row + 3, anchor, last_c)
        _apply_restored_title_band_style(ws, style_ref_col, cost_row, anchor)

    n = 0
    for name, anchor in list(cost_name_to_col.items()):
        ce = find_cost_entry(rate, name)
        if not ce:
            continue
        tit = merged_cell_top_left_value(ws, cost_row, anchor)
        if tit is not None and str(tit).strip():
            continue
        write_band(anchor, ce)
        n += 1

    ordered = _rate_card_ordered_cost_names(rate)
    missing = [x for x in ordered if x not in cost_name_to_col]
    if missing:
        structural = _structural_cost_block_anchors(ws, cost_row, header_row)
        used = set(cost_name_to_col.values())
        free = [a for a in sorted(structural) if a not in used]
        placed = 0
        fi = 0
        for name in missing:
            if fi >= len(free):
                break
            ce = find_cost_entry(rate, name)
            if not ce:
                continue
            anchor = free[fi]
            fi += 1
            cost_name_to_col[name] = anchor
            write_band(anchor, ce)
            n += 1
            placed += 1
        unplaced = [x for x in missing if x not in cost_name_to_col]
        if unplaced:
            report.append(
                f"{len(unplaced)} cost name(s) still have no column anchor "
                f"({placed} placed on free structural column(s))."
            )

    if n:
        report.append(
            f"Restored {n} cost title/condition block(s) from JSON (no green/red highlight)."
        )
        log.debug(
            "restore_unmapped_cost_title_bands_from_rate: filled %s band(s), "
            "%s name→col mapping(s).",
            n,
            len(cost_name_to_col),
        )


def ensure_cost_block_title_band_merges(
    ws,
    cost_row: int,
    header_row: int,
    cost_name_to_col: dict[str, int],
    rate: dict,
    report: list[str],
) -> None:
    """
    Re-merge the fee **title band** (rows ``cost_row`` … ``cost_row + 3``) horizontally.

    **Primary:** use each **Currency** cell on ``header_row`` as the start of a block and
    merge through the column **before** the next **Currency** — “from this Currency until
    the next Currency” (covers AF–AG, AH–AI, AJ–AK, etc.) without relying on name→col or
    JSON MIN alone.

    **Fallback:** if no ``Currency`` markers are found, use ``cost_name_to_col`` +
    :func:`resolve_excel_cost_block_width` per cost (previous behaviour).
    """
    max_c = max(int(ws.max_column or 0), effective_max_column(ws))
    anchors = _lane_currency_anchor_columns(ws, header_row)

    if len(anchors) >= 1:
        n = 0
        for i, anchor in enumerate(anchors):
            if i + 1 < len(anchors):
                last_c = anchors[i + 1] - 1
            else:
                w = infer_excel_cost_block_width(ws, header_row, anchor)
                last_c = min(max_c, anchor + w - 1)
            if last_c < anchor:
                continue
            for r in (cost_row, cost_row + 1, cost_row + 2, cost_row + 3):
                if r >= header_row:
                    continue
                _merge_row_span(ws, r, anchor, last_c)
            n += 1
        if n:
            report.append(
                f"Merged fee title band using Currency-row boundaries ({n} block(s); "
                f"lane header row {header_row})."
            )
        log.debug(
            "ensure_cost_block_title_band_merges: currency anchors=%s (showing up to 24).",
            anchors[:24],
        )
        return

    n = 0
    for _name, anchor in sorted(cost_name_to_col.items(), key=lambda x: x[1]):
        ce = find_cost_entry(rate, _name)
        w = resolve_excel_cost_block_width(ws, cost_row, header_row, anchor, ce)
        last_c = anchor + w - 1
        for r in (cost_row, cost_row + 1, cost_row + 2, cost_row + 3):
            if r >= header_row:
                continue
            _merge_row_span(ws, r, anchor, last_c)
        n += 1
    if n:
        report.append(
            f"Merged fee title band from name→col map (fallback, {n} block(s))."
        )
        log.debug(
            "ensure_cost_block_title_band_merges: fallback refreshed %s block(s).",
            n,
        )


def _rate_card_ordered_cost_names(rate: dict) -> list[str]:
    """
    Single source for Rate card cost order: ``cost_names`` from merge, then any extra
    names seen on lane cost rows (deduplicated, first-seen order).
    """
    rc = rate["sheets"]["Rate card"]
    ordered: list[str] = []
    seen: set[str] = set()
    for nm in rc.get("cost_names") or []:
        s = str(nm).strip()
        if s and s not in seen:
            seen.add(s)
            ordered.append(s)
    for ln in rc.get("lanes") or []:
        for c in ln.get("Costs", []):
            nm = c.get("Cost")
            if not nm:
                continue
            s = str(nm).strip()
            if s not in seen:
                seen.add(s)
                ordered.append(s)
    return ordered


def ensure_excel_cost_columns(
    ws,
    cost_row: int,
    header_row: int,
    cost_name_to_col: dict[str, int],
    rate: dict,
    report: list[str],
    mark_green,
    mark_red,
) -> None:
    ordered = _rate_card_ordered_cost_names(rate)
    initial_keys = set(cost_name_to_col.keys())
    missing = [n for n in ordered if n not in initial_keys]
    log.debug(
        "ensure_excel_cost_columns: template headers=%s merged JSON distinct fees=%s missing=%s",
        len(initial_keys),
        len(ordered),
        len(missing),
    )
    if missing:
        log.debug(
            "ensure_excel_cost_columns: missing fee names (first 50): %s",
            missing[:50],
        )
        uth_on_sheet = sorted(
            [(k, cost_name_to_col[k]) for k in cost_name_to_col if "42UTH" in k],
            key=lambda t: t[1],
        )
        log.debug(
            "ensure_excel_cost_columns: template scan already has these '42UTH' headers (name, col): %s",
            uth_on_sheet[:30],
        )
    ref_col = min(cost_name_to_col.values()) if cost_name_to_col else 11
    cond_r = cost_row + 3
    conditional_template = ws.cell(cond_r, ref_col).value
    insert_excel_cost_columns_alphabetically(
        ws,
        cost_row,
        header_row,
        cost_name_to_col,
        missing,
        rate,
        ref_col,
        report,
        mark_green,
        mark_red,
        conditional_template,
    )


def apply_excel_from_rate(
    template_xlsx: Path,
    out_xlsx: Path,
    rate: dict,
    prior_rate: dict | None,
) -> list[str]:
    """
    Write Rate card lane/cost cells; green-fill changed cells vs prior_rate (if provided).
    """
    log.debug(
        "apply_excel_from_rate: copy template %s → %s (then fill Rate card from merged JSON).",
        template_xlsx,
        out_xlsx,
    )
    report: list[str] = []
    shutil.copy2(template_xlsx, out_xlsx)
    wb = load_workbook(out_xlsx)
    if "Rate card" not in wb.sheetnames:
        report.append("Template has no 'Rate card' sheet.")
        log.debug("apply_excel: abort — no 'Rate card' sheet on template.")
        wb.save(out_xlsx)
        return report

    ws = wb["Rate card"]
    header_row = None
    for r in range(1, 40):
        if ws.cell(r, 1).value == "Lane #":
            header_row = r
            break
    if not header_row:
        report.append("Could not find 'Lane #' header row.")
        log.debug("apply_excel: abort — Lane # header not found in rows 1–39.")
        wb.save(out_xlsx)
        return report

    log.debug(
        "apply_excel: header_row=%s cost_title_scan_row=%s (scan row 11..max for cost names).",
        header_row,
        header_row - 4,
    )
    cost_row = header_row - 4
    cost_name_to_col: dict[str, int] = {}
    # Merge-aware anchors + wide scan (same as ``transform_inputs``): fee titles live on merge
    # top-left only; ``max_column`` can under-report on wide sheets — preserves existing blocks.
    max_scan = max(int(ws.max_column or 0), effective_max_column(ws))
    for c in range(11, max_scan + 1):
        if not merge_anchor_cell(ws, cost_row, c):
            continue
        v = merged_cell_top_left_value(ws, cost_row, c)
        if v and str(v).strip():
            cost_name_to_col[str(v).strip()] = c

    def mark_green(r: int, c: int) -> None:
        cell = ws.cell(r, c)
        cell.fill = GREEN_FILL

    def mark_red(r: int, c: int) -> None:
        cell = ws.cell(r, c)
        cell.fill = RED_FILL

    report.append(
        f"Rate card: {len(cost_name_to_col)} cost title anchor(s) from template "
        f"(merge-aware scan, cols 11–{max_scan})."
    )
    log.debug(
        "apply_excel: %s known cost column(s) from template (merge-aware scan before inserts).",
        len(cost_name_to_col),
    )
    ref_col = min(cost_name_to_col.values()) if cost_name_to_col else 11
    conditional_template = ws.cell(cost_row + 3, ref_col).value
    expand_template_cost_blocks_for_minimum(
        ws,
        cost_row,
        header_row,
        cost_name_to_col,
        rate,
        report,
        mark_green,
        mark_red,
        ref_col,
        conditional_template,
    )
    ensure_excel_cost_columns(
        ws, cost_row, header_row, cost_name_to_col, rate, report, mark_green, mark_red
    )
    log.debug(
        "apply_excel: after ensure_excel_cost_columns — %s cost column(s) mapped name→col.",
        len(cost_name_to_col),
    )

    restore_unmapped_cost_title_bands_from_rate(
        ws, cost_row, header_row, cost_name_to_col, rate, report
    )

    ensure_cost_block_title_band_merges(
        ws, cost_row, header_row, cost_name_to_col, rate, report
    )

    prior_lanes = {}
    if prior_rate:
        for ln in prior_rate.get("sheets", {}).get("Rate card", {}).get("lanes", []):
            prior_lanes[_norm(ln.get("Lane #"))] = ln

    lane_row: dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 2):
        v = ws.cell(r, 1).value
        if v is None or str(v).strip() == "":
            continue
        lane_row[str(v).strip()] = r

    max_r = _last_used_row_col_a(ws, header_row)
    _nl = len(rate["sheets"]["Rate card"]["lanes"])
    log.debug(
        "apply_excel: writing %s lane(s) from merged JSON to sheet rows (green=new/changed vs prior).",
        _nl,
    )
    for lane in rate["sheets"]["Rate card"]["lanes"]:
        ln_id = _norm(lane.get("Lane #"))
        r = lane_row.get(ln_id)
        if not r:
            max_r += 1
            r = max_r
            lane_row[ln_id] = r
            report.append(f"Appended Excel row {r} for lane {ln_id}.")

        hdr_map = [
            (1, "Lane #"),
            (2, "KEY"),
            (3, "Rate Card"),
            (4, "Carrier"),
            (5, "SERVICE"),
            (6, "Service"),
            (7, "Valid from"),
            (8, "Valid to"),
            (9, "Origin Port"),
            (10, "Destination Port"),
        ]
        pl = prior_lanes.get(ln_id, {})
        for col, json_key in hdr_map:
            val = lane.get(json_key)
            if val is None:
                val = ""
            old = pl.get(json_key)
            ws.cell(r, col).value = val
            if _norm(old) != _norm(val):
                mark_green(r, col)

        for cost in lane.get("Costs", []):
            name = cost.get("Cost")
            if not name:
                continue
            c = cost_name_to_col.get(str(name).strip())
            if not c:
                report.append(f"No Excel column for cost {name!r} (lane {ln_id}).")
                continue
            w_blk = resolve_excel_cost_block_width(ws, cost_row, header_row, c, cost)
            cur = _norm(cost.get("Currency"))
            pr = cost.get("Price")
            old_costs = {x.get("Cost"): x for x in pl.get("Costs", [])}
            oc = old_costs.get(name)
            if w_blk == 3:
                ws.cell(r, c).value = cur if cur else None
                mnv = cost.get("Minimum")
                if cost_needs_min_layout(cost):
                    try:
                        fv = float(mnv)
                        ws.cell(r, c + 1).value = int(fv) if fv == int(fv) else fv
                    except (TypeError, ValueError):
                        ws.cell(r, c + 1).value = mnv
                else:
                    ws.cell(r, c + 1).value = None
                ws.cell(r, c + 2).value = pr
                om = oc.get("Minimum") if oc else None
                try:
                    om_n = float(om) if om not in (None, "") else None
                    nm_n = float(mnv) if mnv not in (None, "") else None
                    if om_n is None and nm_n is None:
                        min_same = True
                    elif om_n is not None and nm_n is not None:
                        min_same = om_n == nm_n
                    else:
                        min_same = False
                except (TypeError, ValueError):
                    min_same = _norm(str(om if om is not None else "")) == _norm(
                        str(mnv if mnv is not None else "")
                    )
                changed = (
                    not oc
                    or _norm(oc.get("Currency")) != cur
                    or oc.get("Price") != pr
                    or not min_same
                )
                if changed:
                    mark_green(r, c)
                    mark_green(r, c + 1)
                    mark_green(r, c + 2)
            else:
                ws.cell(r, c).value = cur if cur else None
                ws.cell(r, c + 1).value = pr
                if not oc or _norm(oc.get("Currency")) != cur or oc.get("Price") != pr:
                    mark_green(r, c)
                    mark_green(r, c + 1)

        lane_fee_names = {
            str(x.get("Cost")).strip()
            for x in lane.get("Costs") or []
            if x.get("Cost")
        }
        _clear_excel_lane_row_unused_fee_columns(
            ws,
            r,
            header_row,
            cost_row,
            cost_name_to_col,
            lane_fee_names,
            rate,
        )

    log.debug("apply_excel: saving workbook %s", out_xlsx)
    wb.save(out_xlsx)
    return report


def prompt_pick_csv() -> Path | None:
    files = list_update_csvs()
    if not files:
        print("No CSV files in input/update.")
        return None
    print("\ninput/update CSV files:")
    for i, p in enumerate(files, 1):
        print(f"  [{i}] {p.name}")
    raw = input("Select file number: ").strip()
    try:
        idx = int(raw)
    except ValueError:
        return None
    if 1 <= idx <= len(files):
        return files[idx - 1]
    return None


def prompt_pick_rate_json() -> Path | None:
    files = list_rate_jsons()
    if not files:
        print("No rate JSON files in processing/rate.")
        return None
    print("\nprocessing/rate — baseline JSON to merge into:")
    for i, p in enumerate(files, 1):
        try:
            rel = p.relative_to(ROOT)
        except ValueError:
            rel = p
        print(f"  [{i}] {rel}")
    raw = input("Select file number: ").strip()
    try:
        idx = int(raw)
    except ValueError:
        return None
    if 1 <= idx <= len(files):
        return files[idx - 1]
    return None


def prompt_pick_template_xlsx() -> Path | None:
    """
    Interactive pick for ``input/rate/*.xlsx`` — the workbook used to discover fee column
    layout (see :func:`apply_excel_from_rate`). Should match the same agreement / column
    layout as the baseline rate JSON you already selected.
    """
    files = list_rate_templates()
    if not files:
        print("No template xlsx in input/rate.")
        return None
    print("\ninput/rate — Excel template (fee header row is scanned from this file):")
    for i, p in enumerate(files, 1):
        try:
            rel = p.relative_to(ROOT)
        except ValueError:
            rel = p
        print(f"  [{i}] {rel}")
    raw = input("Select file number: ").strip()
    try:
        idx = int(raw)
    except ValueError:
        return None
    if 1 <= idx <= len(files):
        return files[idx - 1]
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Combine ocean update CSV and merge into rate JSON.")
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print step-by-step DEBUG trace (combine, merge records, Excel apply).",
    )
    parser.add_argument("--csv", type=Path, help="Path to Ocean Rates CSV under input/update.")
    parser.add_argument(
        "--rate-json",
        type=Path,
        help="Rate workbook JSON under processing/rate (default: interactive pick).",
    )
    parser.add_argument(
        "--template-xlsx",
        type=Path,
        help=(
            "Template Excel under input/rate. If omitted: use input/rate/<same stem as baseline "
            "JSON>.xlsx when present; else only file if just one; else interactive pick."
        ),
    )
    args = parser.parse_args()
    configure_update_logging(args.debug)

    csv_path = args.csv
    if not csv_path:
        csv_path = prompt_pick_csv()
    if not csv_path or not csv_path.is_file():
        print("No CSV selected.")
        return
    csv_path = csv_path.resolve()
    print(f"\nUsing CSV: {csv_path}")

    rate_json_path = args.rate_json
    if not rate_json_path:
        rate_json_path = prompt_pick_rate_json()
    if not rate_json_path or not rate_json_path.is_file():
        print("No rate JSON selected or file missing.")
        return
    rate_json_path = rate_json_path.resolve()
    print(f"Using baseline rate JSON: {rate_json_path}\n")

    if args.template_xlsx:
        tpl = args.template_xlsx.resolve()
        if not tpl.is_file():
            print(f"Template not found or not a file: {args.template_xlsx}")
            return
    else:
        txs = list_rate_templates()
        if not txs:
            print("No template xlsx in input/rate.")
            return
        json_stem = rate_json_path.stem
        stem_matches = sorted(p for p in txs if p.stem == json_stem)
        if stem_matches:
            tpl = stem_matches[0].resolve()
            try:
                rel = tpl.relative_to(ROOT)
            except ValueError:
                rel = tpl
            print(f"Using template matching baseline JSON stem: {rel}")
            if len(stem_matches) > 1:
                print(
                    f"(Warning: {len(stem_matches)} templates share stem {json_stem!r}; using first.)"
                )
        elif len(txs) == 1:
            tpl = txs[0].resolve()
            try:
                rel = tpl.relative_to(ROOT)
            except ValueError:
                rel = tpl
            print(f"Using only template in input/rate: {rel}")
            print(
                f"(Note: name differs from baseline JSON stem {json_stem!r} — "
                "prefer adding input/rate/<baseline stem>.xlsx to skip this mismatch.)"
            )
        else:
            picked = prompt_pick_template_xlsx()
            if not picked or not picked.is_file():
                print("No template selected.")
                return
            tpl = picked.resolve()

    OUT_COMBINED_DIR.mkdir(parents=True, exist_ok=True)
    OUT_RESULT_DIR.mkdir(parents=True, exist_ok=True)

    log.debug(
        "Pipeline inputs: csv=%s rate_json=%s template_xlsx=%s",
        csv_path,
        rate_json_path,
        tpl,
    )
    prior_preview = json.loads(rate_json_path.read_text(encoding="utf-8"))
    out_stem = result_stem_for_merge_output(rate_json_path, tpl, prior_preview)
    log.debug(
        "Output stem for result JSON/xlsx/report: %r (from rate JSON filename / General info).",
        out_stem,
    )

    stem = safe_json_name(csv_path.stem)
    combined, c_issues = combine_ocean_rates_csv(csv_path)
    combined_path = OUT_COMBINED_DIR / f"{stem}_combined.json"
    combined_path.write_text(
        json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote {combined_path.relative_to(ROOT)}")
    combine_report_path = OUT_COMBINED_DIR / f"{stem}_combine_report.txt"
    combine_report_path.write_text(
        "\n".join(c_issues) if c_issues else "(no inconsistencies)",
        encoding="utf-8",
    )
    print(f"Wrote {combine_report_path.relative_to(ROOT)}")

    report_lines = list(c_issues)
    report_lines.append("--- merge ---")
    report_lines.append(
        f"Excel template for fee column layout: {tpl.name} "
        "(fee columns discovered from this workbook’s fee title row; usually auto-selected "
        "to match baseline JSON filename stem — columns are not read from JSON coordinates)."
    )
    prior = prior_preview
    update_label = csv_path.stem
    log.debug(
        "Calling merge_rate_with_updates with %s combined record(s), label=%r.",
        len(combined["records"]),
        update_label,
    )
    merged_rate, _, _ = merge_rate_with_updates(
        prior,
        combined["records"],
        update_label,
        report_lines,
    )
    bump_version_string_in_general_info(merged_rate)
    merged_rate["source_file"] = f"{out_stem}.xlsx"
    try:
        merged_rate["source_path"] = str(
            (OUT_RESULT_DIR / f"{out_stem}.xlsx").relative_to(ROOT)
        )
    except ValueError:
        merged_rate["source_path"] = str(OUT_RESULT_DIR / f"{out_stem}.xlsx")

    result_json = OUT_RESULT_DIR / f"{out_stem}.json"
    result_json.write_text(
        json.dumps(merged_rate, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote {result_json.relative_to(ROOT)}")

    result_xlsx = OUT_RESULT_DIR / f"{out_stem}.xlsx"
    log.debug("Writing Excel result from template + merged_rate → %s", result_xlsx)
    x_issues = apply_excel_from_rate(tpl, result_xlsx, merged_rate, prior)
    report_lines.extend(x_issues)
    print(f"Wrote {result_xlsx.relative_to(ROOT)}")

    report_path = OUT_RESULT_DIR / f"{out_stem}_report.txt"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")
    print(f"Wrote {report_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
